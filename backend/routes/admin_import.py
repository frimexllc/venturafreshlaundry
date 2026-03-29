"""Admin Import (CSV/Excel) endpoints."""
import io
import csv
import uuid
import logging
from datetime import datetime, timezone

from fastapi import APIRouter, HTTPException, Depends, Query, UploadFile, File

from database import db
from models import ImportMappingSuggestRequest, ImportMappingConfirmRequest
from auth import get_current_user, require_admin
from utils import (
    ensure_ai_indexes, extract_json_payload, call_ollama,
    normalize_header, set_nested_value, suggest_mapping,
    resolve_or_create_customer_from_row,
    generate_order_number, build_order_times,
)

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api", tags=["Admin Import"])


@router.post("/admin/import")
async def create_import(origen: str = Query("csv"), file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    require_admin(current_user)
    await ensure_ai_indexes()
    filename = file.filename or ""
    content = await file.read()
    rows, headers = [], []

    if filename.lower().endswith(".csv") or origen == "csv":
        text = content.decode("utf-8", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        headers = reader.fieldnames or []
        for idx, row in enumerate(reader):
            if idx >= 500:
                break
            rows.append(row)
    elif filename.lower().endswith(".xlsx") or origen == "excel":
        try:
            import openpyxl
        except Exception:
            raise HTTPException(status_code=400, detail="Excel no soportado sin openpyxl")
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        sheet = workbook.active
        headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        for idx, row_cells in enumerate(sheet.iter_rows(min_row=2), start=0):
            if idx >= 500:
                break
            row = {}
            for col_idx, cell in enumerate(row_cells):
                key = headers[col_idx] if col_idx < len(headers) else f"col_{col_idx}"
                row[key] = cell.value
            rows.append(row)
    else:
        raise HTTPException(status_code=400, detail="Formato no soportado")

    import_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    doc = {"id": import_id, "origen": origen, "estado": "subido", "raw_headers": headers, "raw_sample": rows[:5], "raw_rows": rows, "fecha_importacion": now, "usuario_id": current_user["id"]}
    await db.importaciones_legacy.insert_one(doc)
    return {"import_id": import_id, "campos_detectados": headers, "sample": rows[:5]}


@router.post("/admin/import/{import_id}/mapping/suggest")
async def suggest_import_mapping(import_id: str, data: ImportMappingSuggestRequest, current_user: dict = Depends(get_current_user)):
    require_admin(current_user)
    import_doc = await db.importaciones_legacy.find_one({"id": import_id}, {"_id": 0})
    if not import_doc:
        raise HTTPException(status_code=404, detail="Importación no encontrada")
    mapping = suggest_mapping(data.campos_legacy)
    prompt = f"Devuelve JSON con clave sugerencias (objeto) para mapear campos legacy a ordenes. campos={data.campos_legacy}"
    try:
        raw = call_ollama(prompt)
        payload = extract_json_payload(raw)
        ai_mapping = payload.get("sugerencias")
        if isinstance(ai_mapping, dict):
            mapping = {**mapping, **ai_mapping}
    except Exception:
        pass
    await db.importaciones_legacy.update_one({"id": import_id}, {"$set": {"mapping_sugerido": mapping}})
    return {"sugerencias": mapping}


@router.post("/admin/import/{import_id}/mapping/confirm")
async def confirm_import_mapping(import_id: str, data: ImportMappingConfirmRequest, current_user: dict = Depends(get_current_user)):
    require_admin(current_user)
    import_doc = await db.importaciones_legacy.find_one({"id": import_id}, {"_id": 0})
    if not import_doc:
        raise HTTPException(status_code=404, detail="Importación no encontrada")
    rows = import_doc.get("raw_rows", [])
    mapping = data.mapping_campos
    now = datetime.now(timezone.utc).isoformat()
    created_orders = []
    for row in rows:
        mapped = {}
        for legacy_key, target_path in mapping.items():
            if legacy_key in row and target_path:
                set_nested_value(mapped, target_path, row[legacy_key])
        status_value = mapped.get("estado_actual") or mapped.get("status") or "new"
        order_id = str(uuid.uuid4())
        order_number = mapped.get("order_number") or await generate_order_number()
        customer_id = await resolve_or_create_customer_from_row(mapped)
        order = {
            "id": order_id, "order_number": order_number,
            "customer_id": customer_id, "customer_name": mapped.get("customer_name"),
            "service_type": mapped.get("service_type") or "pickup_delivery",
            "pickup_date": mapped.get("formulario", {}).get("pickup_date"),
            "pickup_time_window": mapped.get("formulario", {}).get("pickup_time_window"),
            "pickup_address": mapped.get("formulario", {}).get("pickup_address"),
            "delivery_address": mapped.get("formulario", {}).get("delivery_address"),
            "estimated_lbs": mapped.get("formulario", {}).get("estimated_lbs"),
            "notes": mapped.get("notes"),
            "gate_code": mapped.get("formulario", {}).get("gate_code"),
            "status": status_value, "estado_actual": status_value,
            "payment_status": "unpaid", "total_amount": None,
            "tiempos": build_order_times(mapped.get("tiempos", {}).get("creacion") or now, status_value),
            "errores_validacion": [], "secciones": [],
            "importada": True, "origen": "import_legacy",
            "qr_token": str(uuid.uuid4()),
            "created_at": mapped.get("tiempos", {}).get("creacion") or now,
            "updated_at": now,
        }
        await db.orders.insert_one(order)
        created_orders.append(order_id)
    await db.importaciones_legacy.update_one({"id": import_id}, {"$set": {"estado": "procesado", "mapping_campos": mapping, "ordenes_creadas_ids": created_orders}})
    return {"ok": True, "ordenes_creadas": len(created_orders)}


@router.post("/admin/import/{import_id}/plan-recuperacion")
async def import_recovery_plan(import_id: str, current_user: dict = Depends(get_current_user)):
    require_admin(current_user)
    import_doc = await db.importaciones_legacy.find_one({"id": import_id}, {"_id": 0})
    if not import_doc:
        raise HTTPException(status_code=404, detail="Importación no encontrada")
    order_ids = import_doc.get("ordenes_creadas_ids", [])
    if not order_ids:
        raise HTTPException(status_code=400, detail="No hay órdenes importadas")
    stale_orders = await db.orders.find({"id": {"$in": order_ids}, "status": {"$nin": ["completed", "cancelled"]}}, {"_id": 0}).to_list(200)
    acciones = [{"order_id": o["id"], "status": "processing"} for o in stale_orders[:50]]
    now = datetime.now(timezone.utc).isoformat()
    propuesta = {
        "id": str(uuid.uuid4()), "tipo": "plan_recuperacion",
        "descripcion": "Plan de recuperación generado desde importación legacy",
        "estado": "pendiente", "impacto_estimado": {"ordenes_afectadas": len(stale_orders)},
        "accion_sugerida": {"type": "plan_recuperacion", "payload": {"acciones": acciones}},
        "nivel_riesgo": "medio", "datos_respaldo": {"import_id": import_id},
        "fecha_generacion": now, "fuente": "importaciones_legacy",
    }
    await db.propuestas_ia.insert_one(propuesta)
    await db.importaciones_legacy.update_one({"id": import_id}, {"$set": {"plan_recuperacion_propuesta_id": propuesta["id"]}})
    return {"ok": True, "propuesta_id": propuesta["id"]}
