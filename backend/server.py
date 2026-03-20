from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, Query, Request, UploadFile, File
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse, HTMLResponse, FileResponse, Response
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import io
import csv
import json
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
import requests
import qrcode
from qrcode.image.svg import SvgImage
import zipfile
import socketio
import base64
import html
import time
import re

from normalization import (
    normalize_email,
    normalize_phone,
    normalize_spaces,
    normalize_name,
    normalize_address,
    normalize_yes_no,
    normalize_preference_dict
)

ROOT_DIR = Path(__file__).parent
app_url = os.environ.get("APP_URL", "")
if not app_url or "preview" in app_url or "localhost" in app_url:
    load_dotenv(ROOT_DIR / '.env', override=False)
BUSINESS_NAME = os.environ.get("BUSINESS_NAME", "Ventura Fresh Laundromat")

# Import AI Assistant
try:
    from ai_assistant import generate_daily_briefing, ai_analyze_business, ai_suggest_actions
    AI_ASSISTANT_ENABLED = True
except ImportError:
    AI_ASSISTANT_ENABLED = False
    logger = logging.getLogger(__name__)
    logger.warning("AI Assistant not available")

# Import notification services
try:
    from notifications import (
        notify_order_created,
        notify_order_status_changed,
        send_email,
        send_sms,
        send_voice_call,
        send_whatsapp,
        send_preferred_notification,
        build_notification_content,
        generate_ai_message,
        detect_language,
        normalize_preferred_contact
    )
    NOTIFICATIONS_ENABLED = True
except ImportError:
    NOTIFICATIONS_ENABLED = False
    logger = logging.getLogger(__name__)
    logger.warning("Notification services not available")

try:
    from routes.public_forms import get_public_forms_router
except ImportError:
    get_public_forms_router = None
    logger = logging.getLogger(__name__)
    logger.warning("Public forms router not available")

try:
    from routes.voice import get_voice_router
except ImportError:
    get_voice_router = None
    logger = logging.getLogger(__name__)
    logger.warning("Voice router not available")

# Import n8n integration
try:
    from n8n_integration import n8n_router, set_database as set_n8n_db
    N8N_ENABLED = True
except ImportError:
    N8N_ENABLED = False
    n8n_router = None
    logger = logging.getLogger(__name__)
    logger.warning("n8n integration not available")

# Import store module
try:
    from store import store_router, set_database as set_store_db, handle_stripe_webhook
    STORE_ENABLED = True
except ImportError:
    STORE_ENABLED = False
    store_router = None
    logger = logging.getLogger(__name__)
    logger.warning("Store module not available")

# Stripe Checkout integration (service orders)
try:
    from emergentintegrations.payments.stripe.checkout import (
        StripeCheckout,
        CheckoutSessionResponse,
        CheckoutStatusResponse,
        CheckoutSessionRequest,
    )
    STRIPE_CHECKOUT_AVAILABLE = True
except ImportError:
    STRIPE_CHECKOUT_AVAILABLE = False

    class CheckoutSessionRequest(BaseModel):
        amount: float
        currency: str
        success_url: str
        cancel_url: str
        metadata: Optional[Dict[str, str]] = None

    class CheckoutSessionResponse(BaseModel):
        url: str = ""
        session_id: str = ""

    class CheckoutStatusResponse(BaseModel):
        status: str = ""
        payment_status: str = ""
        amount_total: int = 0
        currency: str = ""
        metadata: Dict[str, str] = {}

    class StripeCheckout:
        def __init__(self, api_key: str, webhook_url: str):
            self.api_key = api_key
            self.webhook_url = webhook_url

        async def create_checkout_session(self, request: CheckoutSessionRequest):
            raise RuntimeError("Stripe integration not available")

        async def get_checkout_status(self, checkout_session_id: str):
            raise RuntimeError("Stripe integration not available")

        async def handle_webhook(self, payload: bytes, signature: str):
            raise RuntimeError("Stripe integration not available")

# Import blog module
try:
    from blog import blog_router, set_database as set_blog_db
    BLOG_ENABLED = True
except ImportError:
    BLOG_ENABLED = False
    blog_router = None
    logger = logging.getLogger(__name__)
    logger.warning("Blog module not available")

# Import automation engine
try:
    from automation_engine import automation_router, set_database as set_automation_db, set_realtime_emitter
    AUTOMATION_ENABLED = True
except ImportError:
    AUTOMATION_ENABLED = False
    automation_router = None
    set_realtime_emitter = None
    logger = logging.getLogger(__name__)
    logger.warning("Automation engine not available")

# Stripe advanced sync scaffold (disabled by default)
try:
    from stripe_sync_scaffold import stripe_sync_router, set_database as set_stripe_sync_db
    STRIPE_SYNC_SCAFFOLD_ENABLED = True
except ImportError:
    STRIPE_SYNC_SCAFFOLD_ENABLED = False
    stripe_sync_router = None
    logger = logging.getLogger(__name__)
    logger.warning("Stripe sync scaffold not available")

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]
SKIP_SERVER_NOTIFICATIONS = os.environ.get('SKIP_SERVER_NOTIFICATIONS', 'false').lower() == 'true'

# Set database for n8n integration
if N8N_ENABLED:
    set_n8n_db(db)

# Set database for store module
if STORE_ENABLED:
    set_store_db(db)

# Set database for blog module
if BLOG_ENABLED:
    set_blog_db(db)

# Set database for automation engine
if AUTOMATION_ENABLED:
    set_automation_db(db)

# Set database for Stripe sync scaffold
if STRIPE_SYNC_SCAFFOLD_ENABLED:
    set_stripe_sync_db(db)

# JWT Config
JWT_SECRET = os.environ.get('JWT_SECRET', 'ventura-fresh-laundry-secret-key-2024')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

fastapi_app = FastAPI(title="Ventura Fresh Laundry CRM")
app = fastapi_app
sio = socketio.AsyncServer(async_mode="asgi", cors_allowed_origins="*")
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ==================== MODELS ====================

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: Optional[str] = "operator"  # Default to operator for new users

class UserLogin(BaseModel):
    email: EmailStr
    password: str

# Role-based access control constants
ROLE_ADMIN = "admin"
ROLE_OPERATOR = "operator"
VALID_ROLES = [ROLE_ADMIN, ROLE_OPERATOR]

# Permissions by role - what each role can access
ROLE_PERMISSIONS = {
    ROLE_ADMIN: ["all"],  # Admin has access to everything
    ROLE_OPERATOR: [
        "orders:read", "orders:update_status",
        "customers:read",
        "services:read",
        "operator_dashboard"
    ]
}

class UserResponse(BaseModel):
    id: str
    email: str
    name: str
    role: str
    created_at: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str
    user: UserResponse

class CustomerCreate(BaseModel):
    name: str
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    preferred_contact: Optional[str] = "email"
    notes: Optional[str] = None
    membership_plan: Optional[str] = None
    membership_status: Optional[str] = None
    membership_start_date: Optional[str] = None
    preferences_id: Optional[str] = None

class CustomerResponse(BaseModel):
    id: str
    name: Optional[str] = ""
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    preferred_contact: Optional[str] = "email"
    notes: Optional[str] = None
    status: Optional[str] = "active"
    total_orders: Optional[int] = 0
    membership_plan: Optional[str] = None
    membership_status: Optional[str] = None
    membership_start_date: Optional[str] = None
    preferences_id: Optional[str] = None
    created_at: Optional[str] = ""
    updated_at: Optional[str] = ""

class PreferenceCreate(BaseModel):
    customer_id: str
    detergent_type: Optional[str] = "standard"
    water_temperature: Optional[str] = None
    fabric_softener: Optional[str] = None
    folding_style: Optional[str] = "standard"
    hanging_instructions: Optional[str] = None
    allergies: Optional[str] = None
    special_instructions: Optional[str] = None
    pickup_time_preference: Optional[str] = None
    gate_code: Optional[str] = None
    hang_dry_items: Optional[List[str]] = []
    fragrance_preference: Optional[str] = "light"

class CustomerPreferenceUpdate(BaseModel):
    detergent_type: Optional[str] = "standard"
    water_temperature: Optional[str] = None
    fabric_softener: Optional[str] = None
    folding_style: Optional[str] = "standard"
    hanging_instructions: Optional[str] = None
    allergies: Optional[str] = None
    special_instructions: Optional[str] = None
    pickup_time_preference: Optional[str] = None
    gate_code: Optional[str] = None
    hang_dry_items: Optional[List[str]] = []
    fragrance_preference: Optional[str] = "light"

class PreferenceResponse(BaseModel):
    id: str
    customer_id: str
    detergent_type: str
    water_temperature: Optional[str]
    fabric_softener: Optional[str]
    folding_style: str
    hanging_instructions: Optional[str]
    allergies: Optional[str]
    special_instructions: Optional[str]
    pickup_time_preference: Optional[str]
    gate_code: Optional[str]
    hang_dry_items: List[str]
    fragrance_preference: str
    version: int
    created_at: str
    updated_at: str

class OrderCreate(BaseModel):
    customer_id: str
    service_type: str  # pickup_delivery, wash_fold, self_service
    pickup_date: Optional[str] = None
    pickup_time_window: Optional[str] = None
    pickup_address: Optional[str] = None
    delivery_address: Optional[str] = None
    estimated_lbs: Optional[float] = None
    notes: Optional[str] = None
    gate_code: Optional[str] = None

class OrderResponse(BaseModel):
    id: str
    order_number: Optional[str] = None
    customer_id: Optional[str] = None
    customer_name: Optional[str] = None
    service_type: Optional[str] = "general"
    pickup_date: Optional[str] = None
    pickup_time_window: Optional[str] = None
    pickup_address: Optional[str] = None
    delivery_address: Optional[str] = None
    estimated_lbs: Optional[float] = None
    actual_lbs: Optional[float] = None
    notes: Optional[str] = None
    preferred_contact: Optional[str] = None
    gate_code: Optional[str] = None
    preferences_id: Optional[str] = None
    preferences_snapshot: Optional[dict] = None
    status: str = "new"
    payment_status: Optional[str] = "pending"
    payment_method: Optional[str] = None
    amount_paid: Optional[float] = None
    change_due: Optional[float] = None
    paid_at: Optional[str] = None
    total_amount: Optional[float] = None
    created_at: Optional[str] = ""
    updated_at: Optional[str] = ""

class OrderPaymentUpdate(BaseModel):
    payment_method: str
    amount_received: Optional[float] = None

class OrderStripeCheckoutRequest(BaseModel):
    origin_url: str

class OrderStripeCheckoutResponse(BaseModel):
    session_id: str
    url: str
    amount: float
    currency: str


class QuoteCreate(BaseModel):
    company_name: str
    contact_name: str
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    industry: Optional[str] = None
    estimated_lbs_per_week: Optional[float] = None
    service_needs: Optional[str] = None
    notes: Optional[str] = None

class QuoteResponse(BaseModel):
    id: str
    quote_number: str
    company_name: str
    contact_name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    industry: Optional[str] = None
    estimated_lbs_per_week: Optional[float] = None
    service_needs: Optional[str] = None
    notes: Optional[str] = None
    status: str
    assigned_to: Optional[str] = None
    follow_up_date: Optional[str] = None
    created_at: str
    updated_at: str

class LeadCreate(BaseModel):
    name: str
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    source: Optional[str] = "website"
    interest_type: Optional[str] = None
    notes: Optional[str] = None

class LeadResponse(BaseModel):
    id: str
    name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    source: str
    interest_type: Optional[str] = None
    notes: Optional[str] = None
    status: str
    converted_to_customer_id: Optional[str] = None
    created_at: str
    updated_at: str

class TicketCreate(BaseModel):
    customer_id: Optional[str] = None
    subject: str
    description: str
    category: Optional[str] = "general"

class TicketResponse(BaseModel):
    id: str
    ticket_number: str
    customer_id: Optional[str] = None
    customer_name: Optional[str] = None
    subject: str
    description: str
    category: str
    priority: str
    status: str
    assigned_to: Optional[str] = None
    resolution: Optional[str] = None
    created_at: str
    updated_at: str

class ServiceCreate(BaseModel):
    name: str
    category: Optional[str] = None
    description: Optional[str] = None
    price: Optional[float] = None
    price_unit: Optional[str] = None
    is_active: bool = True
    sort_order: Optional[int] = 0

class ServiceResponse(BaseModel):
    id: str
    name: str
    category: Optional[str] = None
    description: Optional[str] = None
    price: Optional[float] = None
    price_unit: Optional[str] = None
    is_active: bool
    sort_order: Optional[int] = 0
    created_at: str
    updated_at: str

class MembershipSectionUpdate(BaseModel):
    heading: str
    subheading: Optional[str] = None
    special_title: Optional[str] = None
    special_text: Optional[str] = None
    cta_title: Optional[str] = None
    cta_text: Optional[str] = None
    cta_button_label: Optional[str] = None
    cta_button_url: Optional[str] = None
    contact_phone: Optional[str] = None
    is_active: bool = True

class MembershipSectionResponse(BaseModel):
    id: str
    heading: str
    subheading: Optional[str] = None
    special_title: Optional[str] = None
    special_text: Optional[str] = None
    cta_title: Optional[str] = None
    cta_text: Optional[str] = None
    cta_button_label: Optional[str] = None
    cta_button_url: Optional[str] = None
    contact_phone: Optional[str] = None
    is_active: bool
    created_at: str
    updated_at: str

class MembershipPlanCreate(BaseModel):
    name: str
    price: str
    image_url: Optional[str] = None
    features: List[str]
    is_popular: bool = False
    is_active: bool = True
    sort_order: Optional[int] = 0

class MembershipPlanResponse(BaseModel):
    id: str
    name: str
    price: str
    image_url: Optional[str] = None
    features: List[str]
    is_popular: bool
    is_active: bool
    sort_order: Optional[int] = 0
    created_at: str
    updated_at: str

class MembershipSignupResponse(BaseModel):
    id: str
    first_name: Optional[str] = ""
    last_name: Optional[str] = ""
    email: Optional[str] = ""
    phone: Optional[str] = ""
    contact_method: Optional[str] = ""
    address_line1: Optional[str] = ""
    address_line2: Optional[str] = None
    city: Optional[str] = ""
    state: Optional[str] = ""
    zip_code: Optional[str] = ""
    membership_plan: Optional[str] = ""
    plan_name: Optional[str] = None
    plan_id: Optional[str] = None
    laundry_frequency: Optional[str] = ""
    estimated_lbs: Optional[float] = 0
    amount: Optional[float] = None
    payment_status: Optional[str] = None
    status: str = "pending"
    customer_id: Optional[str] = None
    preferences: Optional[dict] = None
    customer_name: Optional[str] = None
    customer_email: Optional[str] = None
    customer_phone: Optional[str] = None
    stripe_session_id: Optional[str] = None
    created_at: Optional[str] = ""
    updated_at: Optional[str] = ""

class MembershipSignupUpdate(BaseModel):
    status: Optional[str] = None
    customer_id: Optional[str] = None

class MembershipCustomerUpdate(BaseModel):
    membership_plan: Optional[str] = None
    membership_status: Optional[str] = None
    membership_start_date: Optional[str] = None

class AdminAIRequest(BaseModel):
    message: str
    execute: bool = True
    session_id: Optional[str] = None
    confirm_token: Optional[str] = None
    context: Optional[dict] = None

class AdminAIInsightsRequest(BaseModel):
    type: str

class PatternScanRequest(BaseModel):
    periodo_desde: Optional[str] = None
    periodo_hasta: Optional[str] = None
    scope: Optional[str] = "orders"
    filtros: Optional[dict] = None

class ProposalGenerateRequest(BaseModel):
    patrones_ids: Optional[List[str]] = None
    max_propuestas: Optional[int] = 10

class ProposalActionRequest(BaseModel):
    accion: str
    modificaciones: Optional[dict] = None
    comentarios: Optional[str] = None

class ImportMappingSuggestRequest(BaseModel):
    campos_legacy: List[str]

class ImportMappingConfirmRequest(BaseModel):
    mapping_campos: Dict[str, str]

class RulesUpdateRequest(BaseModel):
    rules: Dict[str, Any]

class QrResolveRequest(BaseModel):
    qr_token: Optional[str] = None
    payload: Optional[str] = None

class IngestCreate(BaseModel):
    source_form: str
    data: dict

class AuditLogResponse(BaseModel):
    id: str
    event_type: str
    entity_type: str
    entity_id: str
    user_id: Optional[str] = None
    details: Optional[dict] = None
    created_at: str

class DashboardStats(BaseModel):
    total_customers: int
    total_orders: int
    pending_orders: int
    open_tickets: int
    active_quotes: int
    new_leads: int
    orders_today: int
    revenue_this_month: float

# ==================== AUTH HELPERS ====================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str, email: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("sub")
        if not user_id:
            raise HTTPException(status_code=401, detail="Invalid token")
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

def create_customer_token(customer_id: str, email: str) -> str:
    payload = {
        "sub": customer_id,
        "email": email,
        "type": "customer",
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS * 7)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_customer(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        customer_id = payload.get("sub")
        token_type = payload.get("type")
        if not customer_id or token_type != "customer":
            raise HTTPException(status_code=401, detail="Invalid token")
        customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
        if not customer:
            raise HTTPException(status_code=401, detail="Customer not found")
        return customer
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def emit_realtime(event: str, payload: dict):
    try:
        await sio.emit(event, payload)
    except Exception as exc:
        logger.warning(f"Realtime emit failed: {exc}")

if AUTOMATION_ENABLED and set_realtime_emitter:
    set_realtime_emitter(emit_realtime)


def require_admin(current_user: dict):
    if current_user.get("role") != ROLE_ADMIN:
        raise HTTPException(status_code=403, detail="Admin access required")

def require_role(allowed_roles: List[str]):
    """Dependency factory that creates a role checker"""
    def checker(current_user: dict = Depends(get_current_user)):
        user_role = current_user.get("role", ROLE_OPERATOR)
        if user_role == ROLE_ADMIN:  # Admin always passes
            return current_user
        if user_role not in allowed_roles:
            raise HTTPException(
                status_code=403, 
                detail=f"Access denied. Required roles: {allowed_roles}"
            )
        return current_user
    return checker

def has_permission(current_user: dict, permission: str) -> bool:
    """Check if user has specific permission"""
    user_role = current_user.get("role", ROLE_OPERATOR)
    if user_role == ROLE_ADMIN:
        return True
    permissions = ROLE_PERMISSIONS.get(user_role, [])
    return "all" in permissions or permission in permissions

def require_permission(permission: str):
    """Dependency factory for permission-based access"""
    def checker(current_user: dict = Depends(get_current_user)):
        if not has_permission(current_user, permission):
            raise HTTPException(
                status_code=403,
                detail=f"Permission denied: {permission}"
            )
        return current_user
    return checker

def extract_json_payload(text: str):
    cleaned = text.strip()
    if "```" in cleaned:
        start = cleaned.find("```")
        end = cleaned.rfind("```")
        if end > start:
            cleaned = cleaned[start + 3:end].strip()
            if cleaned.startswith("json"):
                cleaned = cleaned[4:].strip()
    return json.loads(cleaned)

def call_ollama(prompt: str):
    """Use Groq API for AI responses - faster than local Ollama"""
    import os
    from groq import Groq
    
    api_key = os.environ.get("GROQ_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="Groq API key not configured")
    
    client = Groq(api_key=api_key)
    models = ["llama-3.3-70b-versatile", "llama-3.1-8b-instant"]
    last_error = None

    for model in models:
        for attempt in range(3):
            try:
                chat_completion = client.chat.completions.create(
                    messages=[
                        {"role": "user", "content": prompt}
                    ],
                    model=model,
                    temperature=0.65,
                    max_tokens=2048
                )
                return chat_completion.choices[0].message.content.strip()
            except Exception as e:
                last_error = e
                wait_seconds = 0.6 * (attempt + 1)
                logger.warning(f"Groq API retry model={model} attempt={attempt + 1}: {e}")
                time.sleep(wait_seconds)
                continue

    logger.error(f"Groq API error after retries: {last_error}")
    raise HTTPException(status_code=502, detail=f"AI service error: {str(last_error)}")

ai_indexes_ready = False

async def ensure_ai_indexes():
    global ai_indexes_ready
    if ai_indexes_ready:
        return
    await db.patrones_detectados.create_index([("fecha_deteccion", -1)])
    await db.propuestas_ia.create_index([("estado", 1), ("fecha_generacion", -1)])
    await db.importaciones_legacy.create_index([("origen", 1), ("fecha_importacion", -1)])
    await db.audit_logs.create_index([("created_at", -1)])
    ai_indexes_ready = True

async def get_or_seed_business_rules():
    rules = await db.reglas_negocio.find_one({"id": "order_rules_v1"}, {"_id": 0})
    if rules:
        return rules
    now = datetime.now(timezone.utc).isoformat()
    rules = {
        "id": "order_rules_v1",
        "type": "order_rules",
        "auto_transitions": {
            "pickup_delivery": {"notify_status": "out_for_delivery"},
            "wash_fold": {"notify_status": "ready"},
            "self_service": {"notify_status": "ready"}
        },
        "sla_hours": {
            "pickup_delivery": 48,
            "wash_fold": 36,
            "self_service": 24
        },
        "created_at": now,
        "updated_at": now
    }
    await db.reglas_negocio.insert_one(rules)
    return rules

def build_order_times(now_iso: str, status_value: str):
    return {
        "creacion": now_iso,
        "ultimo_cambio_estado": now_iso,
        "fechas_estado": {status_value: now_iso}
    }

def validate_order_payload(data: OrderCreate):
    errors = []
    if data.service_type == "pickup_delivery":
        if not data.pickup_date:
            errors.append({"codigo": "MISSING_PICKUP_DATE", "campo": "pickup_date"})
        if not data.pickup_time_window:
            errors.append({"codigo": "MISSING_PICKUP_TIME", "campo": "pickup_time_window"})
        if not data.pickup_address:
            errors.append({"codigo": "MISSING_PICKUP_ADDRESS", "campo": "pickup_address"})
    return errors

def normalize_header(value: str):
    return value.strip().lower()

def set_nested_value(target: dict, path: str, value):
    parts = path.split(".")
    current = target
    for key in parts[:-1]:
        if key not in current or not isinstance(current[key], dict):
            current[key] = {}
        current = current[key]
    current[parts[-1]] = value

def build_qr_svg(payload: str):
    img = qrcode.make(payload, image_factory=SvgImage, box_size=10, border=2)
    buffer = io.BytesIO()
    img.save(buffer)
    return buffer.getvalue()

def build_qr_payload(order: dict):
    return json.dumps({
        "order_id": order.get("id"),
        "order_number": order.get("order_number"),
        "qr_token": order.get("qr_token")
    })


def build_display_order_number(order: dict) -> str:
    order_number = order.get("order_number")
    if order_number and order_number.startswith("VFL-"):
        return order_number
    created_at = order.get("created_at") or datetime.now(timezone.utc).isoformat()
    date_part = created_at[:10].replace("-", "")
    base_id = order_number or order.get("id") or "00000000"
    short = "".join([c for c in str(base_id) if c.isalnum()]).lower()[:8]
    if len(short) < 8:
        short = (short + "00000000")[:8]
    return f"VFL-{date_part}-{short}"


def format_time_window(window: Optional[str]) -> str:
    if not window:
        return "-"
    cleaned = window.replace(" ", "")
    if "-" in cleaned:
        start, end = cleaned.split("-", 1)
        return f"{start} - {end}"
    return window


def build_ticket_lines(order: dict, customer: Optional[dict]) -> List[str]:
    customer = customer or {}
    display_id = build_display_order_number(order)
    status = normalize_status(order.get("status") or "new").upper()
    name = order.get("customer_name") or customer.get("name") or "-"
    phone = customer.get("phone") or order.get("customer_phone") or "-"
    contact = customer.get("preferred_contact") or order.get("preferred_contact") or "-"
    contact_label = str(contact).capitalize() if contact else "-"
    pickup_date = order.get("pickup_date") or "-"
    window = format_time_window(order.get("pickup_time_window") or order.get("pickup_time"))
    address = order.get("pickup_address") or order.get("delivery_address") or customer.get("address") or "-"
    membership = "yes" if customer.get("membership_plan") or customer.get("membership_status") else "no"
    notes = order.get("notes") or "N/A"
    def format_lbs(value):
        if value is None or value == "":
            return "N/A"
        try:
            return f"{float(value):g}"
        except Exception:
            return str(value)
    est_lbs = format_lbs(order.get("estimated_lbs"))
    act_lbs = format_lbs(order.get("actual_lbs"))
    pref_id = order.get("preferences_id") or "N/A"
    customer_id = order.get("customer_id") or customer.get("id") or "N/A"
    email = customer.get("email") or order.get("customer_email") or ""
    source = order.get("origen") or "crm"
    dedup = f"e:{email}|f:{source}"
    if len(dedup) > 45:
        dedup = dedup[:42] + "..."
    summary = f"{display_id} | {pickup_date} {window} | {name}"
    return [
        "VENTURA FRESH LAUNDRY",
        f"ORDER: {display_id}",
        f"STATUS: {status or 'NEW'}",
        f"NAME: {name}",
        f"PHONE: {phone}    CONTACT: {contact_label}",
        f"PICKUP: {pickup_date}    WINDOW: {window}",
        f"ADDR: {address}",
        f"MEMBERSHIP: {membership}",
        f"NOTES: {notes}",
        f"EST_LBS: {est_lbs}",
        f"ACT_LBS: {act_lbs}",
        f"PREF: {pref_id}",
        f"CUS_ID: {customer_id}",
        f"DEDUP: {dedup}",
        "",
        summary
    ]


def build_qr_png_base64(payload: str) -> str:
    qr = qrcode.QRCode(box_size=6, border=2)
    qr.add_data(payload)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    return base64.b64encode(buffer.getvalue()).decode("utf-8")


def build_ticket_svg(order: dict, customer: Optional[dict], qr_payload: str) -> bytes:
    lines = build_ticket_lines(order, customer)
    qr_base64 = build_qr_png_base64(qr_payload)
    qr_size = 180
    padding = 20
    line_height = 16
    font_size = 12
    text_x = padding + qr_size + 20
    height = max(qr_size + padding * 2, padding * 2 + line_height * len(lines))
    width = 760

    text_lines = []
    for index, line in enumerate(lines):
        dy = line_height if index > 0 else 0
        text_lines.append(
            f"<tspan x='{text_x}' dy='{dy}'>{html.escape(line)}</tspan>"
        )

    svg = f"""
<svg xmlns='http://www.w3.org/2000/svg' width='{width}' height='{height}'>
  <rect width='100%' height='100%' fill='white'/>
  <image href='data:image/png;base64,{qr_base64}' x='{padding}' y='{padding}' width='{qr_size}' height='{qr_size}' />
  <text x='{text_x}' y='{padding + font_size}' font-family='Courier New, monospace' font-size='{font_size}' fill='#111'>
    {''.join(text_lines)}
  </text>
</svg>
"""
    return svg.encode("utf-8")

def parse_qr_payload(payload: str):
    try:
        data = json.loads(payload)
        if isinstance(data, dict):
            return data
    except Exception:
        return {}
    return {}

def build_address_parts(address: Optional[str]):
    if not address:
        return {"full": None, "street": None, "city": None, "postal_code": None}
    parts = [p.strip() for p in address.split(",") if p.strip()]
    street = parts[0] if parts else address
    city = parts[1] if len(parts) > 1 else None
    postal_code = None
    if len(parts) > 2:
        postal_code = parts[-1].split()[-1]
    return {"full": address, "street": street, "city": city, "postal_code": postal_code}

def suggest_mapping(headers: List[str]):
    mapping = {}
    for header in headers:
        key = normalize_header(header)
        if key in ["issue key", "key", "id", "order_number", "order number"]:
            mapping[header] = "order_number"
        elif key in ["status", "estado", "state"]:
            mapping[header] = "estado_actual"
        elif key in ["created", "created_at", "fecha", "creation date"]:
            mapping[header] = "tiempos.creacion"
        elif key in ["customer", "customer_name", "name", "cliente"]:
            mapping[header] = "customer_name"
        elif key in ["email", "correo", "customer_email"]:
            mapping[header] = "customer_email"
        elif key in ["phone", "telefono", "customer_phone"]:
            mapping[header] = "customer_phone"
        elif key in ["service_type", "service", "tipo servicio"]:
            mapping[header] = "service_type"
        elif key in ["notes", "summary", "descripcion", "description"]:
            mapping[header] = "notes"
    return mapping

async def resolve_or_create_customer_from_row(row: dict):
    email = row.get("customer_email")
    phone = row.get("customer_phone")
    name = row.get("customer_name") or "Legacy"
    query = []
    if email:
        query.append({"email": email.lower()})
    if phone:
        query.append({"phone": phone})
    if query:
        existing = await db.customers.find_one({"$or": query}, {"_id": 0})
        if existing:
            return existing["id"]
    customer_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    customer = {
        "id": customer_id,
        "name": name,
        "email": email.lower() if email else None,
        "phone": phone,
        "address": None,
        "preferred_contact": "email",
        "notes": "Importación legacy",
        "status": "active",
        "total_orders": 0,
        "created_at": now,
        "updated_at": now
    }
    await db.customers.insert_one(customer)
    return customer_id

# ==================== AUDIT LOG HELPER ====================

async def create_audit_log(event_type: str, entity_type: str, entity_id: str, user_id: str = None, details: dict = None):
    log = {
        "id": str(uuid.uuid4()),
        "event_type": event_type,
        "entity_type": entity_type,
        "entity_id": entity_id,
        "user_id": user_id,
        "details": details,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.audit_logs.insert_one(log)

# ==================== AUTH ENDPOINTS ====================

@api_router.post("/auth/register", response_model=TokenResponse)
async def register(user_data: UserCreate):
    existing = await db.users.find_one({"email": user_data.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Check if this is the first user - make them admin
    user_count = await db.users.count_documents({})
    role = ROLE_ADMIN if user_count == 0 else ROLE_OPERATOR
    
    user_id = str(uuid.uuid4())
    user = {
        "id": user_id,
        "email": user_data.email.lower(),
        "password_hash": hash_password(user_data.password),
        "name": user_data.name,
        "role": role,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user)
    await create_audit_log("USER_CREATED", "user", user_id, user_id)
    
    token = create_token(user_id, user["email"])
    return TokenResponse(
        access_token=token,
        token_type="bearer",
        user=UserResponse(id=user_id, email=user["email"], name=user["name"], role=user["role"], created_at=user["created_at"])
    )

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email.lower()}, {"_id": 0})
    if not user or not verify_password(credentials.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    token = create_token(user["id"], user["email"])
    return TokenResponse(
        access_token=token,
        token_type="bearer",
        user=UserResponse(id=user["id"], email=user["email"], name=user["name"], role=user["role"], created_at=user["created_at"])
    )

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user: dict = Depends(get_current_user)):
    return UserResponse(
        id=current_user["id"],
        email=current_user["email"],
        name=current_user["name"],
        role=current_user["role"],
        created_at=current_user["created_at"]
    )

# ==================== DASHBOARD ====================

@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
    month_start = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
    
    total_customers = await db.customers.count_documents({})
    total_orders = await db.orders.count_documents({})
    pending_orders = await db.orders.count_documents({"status": {"$in": ["new", "processing", "ready"]}})
    open_tickets = await db.tickets.count_documents({"status": {"$in": ["open", "in_progress"]}})
    active_quotes = await db.quotes.count_documents({"status": {"$in": ["new", "sent", "negotiating"]}})
    new_leads = await db.leads.count_documents({"status": "new"})
    orders_today = await db.orders.count_documents({"created_at": {"$gte": today}})
    
    # Calculate revenue
    pipeline = [
        {"$match": {"created_at": {"$gte": month_start}, "payment_status": "paid"}},
        {"$group": {"_id": None, "total": {"$sum": "$total_amount"}}}
    ]
    revenue_result = await db.orders.aggregate(pipeline).to_list(1)
    revenue = revenue_result[0]["total"] if revenue_result else 0
    
    return DashboardStats(
        total_customers=total_customers,
        total_orders=total_orders,
        pending_orders=pending_orders,
        open_tickets=open_tickets,
        active_quotes=active_quotes,
        new_leads=new_leads,
        orders_today=orders_today,
        revenue_this_month=revenue or 0
    )

@api_router.get("/dashboard/recent-activity")
async def get_recent_activity(current_user: dict = Depends(get_current_user)):
    logs = await db.audit_logs.find({}, {"_id": 0}).sort("created_at", -1).limit(20).to_list(20)
    return logs

# ==================== CUSTOMERS ====================

@api_router.post("/customers", response_model=CustomerResponse)
async def create_customer(data: CustomerCreate, current_user: dict = Depends(get_current_user)):
    customer_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    normalized_name = normalize_name(data.name)
    normalized_email = normalize_email(data.email) if data.email else ""
    normalized_phone = normalize_phone(data.phone)
    normalized_address = normalize_address(data.address)
    customer = {
        "id": customer_id,
        "name": normalized_name or data.name,
        "email": normalized_email or (data.email.lower() if data.email else None),
        "phone": normalized_phone or data.phone,
        "address": normalized_address or data.address,
        "preferred_contact": normalize_spaces(data.preferred_contact) or data.preferred_contact,
        "notes": normalize_spaces(data.notes),
        "status": "active",
        "total_orders": 0,
        "membership_plan": normalize_spaces(data.membership_plan),
        "membership_status": normalize_spaces(data.membership_status),
        "membership_start_date": data.membership_start_date,
        "preferences_id": data.preferences_id,
        "created_at": now,
        "updated_at": now
    }
    await db.customers.insert_one(customer)
    await create_audit_log("CUSTOMER_CREATED", "customer", customer_id, current_user["id"])
    customer.pop("_id", None)
    return CustomerResponse(**customer)

@api_router.get("/customers", response_model=List[CustomerResponse])
async def get_customers(
    search: Optional[str] = None,
    status: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=100),
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}}
        ]
    if status:
        query["status"] = status
    
    skip = (page - 1) * page_size
    customers = await db.customers.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    return [CustomerResponse(**c) for c in customers]

@api_router.get("/customers/{customer_id}", response_model=CustomerResponse)
async def get_customer(customer_id: str, current_user: dict = Depends(get_current_user)):
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    return CustomerResponse(**customer)

@api_router.put("/customers/{customer_id}", response_model=CustomerResponse)
async def update_customer(customer_id: str, data: CustomerCreate, current_user: dict = Depends(get_current_user)):
    update_data = data.model_dump(exclude_unset=True)
    if "name" in update_data:
        update_data["name"] = normalize_name(update_data["name"]) or update_data["name"]
    if "email" in update_data and update_data["email"]:
        normalized_email = normalize_email(update_data["email"])
        update_data["email"] = normalized_email or update_data["email"].lower()
    if "phone" in update_data:
        normalized_phone = normalize_phone(update_data["phone"])
        update_data["phone"] = normalized_phone or update_data["phone"]
    if "address" in update_data:
        update_data["address"] = normalize_address(update_data["address"]) or update_data["address"]
    if "preferred_contact" in update_data:
        update_data["preferred_contact"] = normalize_spaces(update_data["preferred_contact"]) or update_data["preferred_contact"]
    if "notes" in update_data:
        update_data["notes"] = normalize_spaces(update_data["notes"])
    if "membership_plan" in update_data:
        update_data["membership_plan"] = normalize_spaces(update_data["membership_plan"])
    if "membership_status" in update_data:
        update_data["membership_status"] = normalize_spaces(update_data["membership_status"])

    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.customers.update_one({"id": customer_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    await create_audit_log("CUSTOMER_UPDATED", "customer", customer_id, current_user["id"])
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    return CustomerResponse(**customer)

@api_router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.customers.delete_one({"id": customer_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    await create_audit_log("CUSTOMER_DELETED", "customer", customer_id, current_user["id"])
    return {"message": "Customer deleted"}

# ==================== PREFERENCES ====================

def normalize_preference_payload(data: PreferenceCreate) -> Dict[str, Any]:
    def normalize_list(value):
        if not value:
            return []
        if isinstance(value, list):
            return [normalize_spaces(v) for v in value if normalize_spaces(v)]
        if isinstance(value, str):
            cleaned = normalize_spaces(value)
            return [v for v in (item.strip() for item in cleaned.split(",")) if v]
        return []

    return {
        "detergent_type": normalize_spaces(data.detergent_type) or "standard",
        "water_temperature": normalize_spaces(data.water_temperature),
        "fabric_softener": normalize_spaces(data.fabric_softener),
        "folding_style": normalize_spaces(data.folding_style) or "standard",
        "hanging_instructions": normalize_spaces(data.hanging_instructions),
        "allergies": normalize_spaces(data.allergies),
        "special_instructions": normalize_spaces(data.special_instructions),
        "pickup_time_preference": normalize_spaces(data.pickup_time_preference),
        "gate_code": normalize_spaces(data.gate_code),
        "hang_dry_items": normalize_list(data.hang_dry_items),
        "fragrance_preference": normalize_spaces(data.fragrance_preference) or "light"
    }

@api_router.post("/preferences", response_model=PreferenceResponse)
async def create_preference(data: PreferenceCreate, current_user: dict = Depends(get_current_user)):
    existing = await db.preferences.find_one({"customer_id": data.customer_id}, sort=[("version", -1)], projection={"_id": 0, "version": 1})
    version = ((existing or {}).get("version", 0) + 1)

    pref_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    normalized = normalize_preference_payload(data)
    pref = {
        "id": pref_id,
        "customer_id": data.customer_id,
        **normalized,
        "version": version,
        "created_at": now,
        "updated_at": now
    }
    await db.preferences.insert_one(pref)
    await db.customers.update_one({"id": data.customer_id}, {"$set": {"preferences_id": pref_id, "updated_at": now}})
    await create_audit_log("PREFERENCE_CREATED", "preference", pref_id, current_user["id"])
    return PreferenceResponse(**pref)

@api_router.get("/preferences/customer/{customer_id}", response_model=PreferenceResponse)
async def get_customer_preference(customer_id: str, current_user: dict = Depends(get_current_user)):
    pref = await db.preferences.find({"customer_id": customer_id}, {"_id": 0}).sort("version", -1).limit(1).to_list(1)
    if not pref:
        raise HTTPException(status_code=404, detail="Preferences not found")
    return PreferenceResponse(**pref[0])

@api_router.get("/customer/preferences", response_model=PreferenceResponse)
async def get_current_customer_preferences(current_customer: dict = Depends(get_current_customer)):
    pref = await db.preferences.find({"customer_id": current_customer["id"]}, {"_id": 0}).sort("version", -1).limit(1).to_list(1)
    if not pref:
        raise HTTPException(status_code=404, detail="Preferences not found")
    return PreferenceResponse(**pref[0])

@api_router.post("/customer/preferences", response_model=PreferenceResponse)
async def upsert_customer_preferences(data: CustomerPreferenceUpdate, current_customer: dict = Depends(get_current_customer)):
    existing = await db.preferences.find({"customer_id": current_customer["id"]}).sort("version", -1).limit(1).to_list(1)
    version = (existing[0]["version"] + 1) if existing else 1

    pref_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    normalized_data = normalize_preference_dict(data.model_dump())
    normalized = normalize_preference_payload(PreferenceCreate(customer_id=current_customer["id"], **(normalized_data or {})))
    pref = {
        "id": pref_id,
        "customer_id": current_customer["id"],
        **normalized,
        "version": version,
        "created_at": now,
        "updated_at": now
    }
    await db.preferences.insert_one(pref)
    await db.customers.update_one({"id": current_customer["id"]}, {"$set": {"preferences_id": pref_id, "updated_at": now}})
    return PreferenceResponse(**pref)

@api_router.delete("/customer/preferences")
async def delete_customer_preferences(current_customer: dict = Depends(get_current_customer)):
    result = await db.preferences.delete_many({"customer_id": current_customer["id"]})
    await db.customers.update_one({"id": current_customer["id"]}, {"$set": {"preferences_id": None, "updated_at": datetime.now(timezone.utc).isoformat()}})
    return {"message": f"Deleted {result.deleted_count} preferences"}

# ==================== ORDERS ====================

async def generate_order_number():
    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    unique = uuid.uuid4().hex[:8]
    return f"VFL-{today}-{unique}"

@api_router.post("/orders", response_model=OrderResponse)
async def create_order(data: OrderCreate, notify: bool = True, current_user: dict = Depends(get_current_user)):
    customer = await db.customers.find_one({"id": data.customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    order_id = str(uuid.uuid4())
    order_number = await generate_order_number()
    now = datetime.now(timezone.utc).isoformat()
    errors = validate_order_payload(data)

    normalized_pickup_address = normalize_address(data.pickup_address or customer.get("address"))
    normalized_delivery_address = normalize_address(data.delivery_address)
    normalized_notes = normalize_spaces(data.notes)
    normalized_gate_code = normalize_spaces(data.gate_code)
    normalized_service_type = normalize_spaces(data.service_type).lower().replace(" ", "_")

    pref = await db.preferences.find({"customer_id": data.customer_id}, {"_id": 0}).sort("version", -1).limit(1).to_list(1)
    preference_id = pref[0].get("id") if pref else None
    preference_snapshot = None
    if pref:
        preference_snapshot = {k: v for k, v in pref[0].items() if k not in ["_id", "customer_id"]}

    order = {
        "id": order_id,
        "order_number": order_number,
        "customer_id": data.customer_id,
        "customer_name": customer["name"],
        "service_type": normalized_service_type,
        "pickup_date": data.pickup_date,
        "pickup_time_window": data.pickup_time_window,
        "pickup_address": normalized_pickup_address,
        "delivery_address": normalized_delivery_address,
        "estimated_lbs": data.estimated_lbs,
        "actual_lbs": None,
        "notes": normalized_notes,
        "gate_code": normalized_gate_code,
        "preferences_id": preference_id,
        "preferences_snapshot": preference_snapshot,
        "status": "new",
        "estado_actual": "new",
        "payment_status": "unpaid",
        "total_amount": None,
        "tiempos": build_order_times(now, "new"),
        "errores_validacion": [
            {**error, "mensaje": error["codigo"], "timestamp": now}
            for error in errors
        ],
        "secciones": [
            {"nombre": "ingesta", "estado": "done", "inicio": now, "fin": now, "errores": []},
            {"nombre": "procesamiento", "estado": "pending", "inicio": None, "fin": None, "errores": []}
        ],
        "importada": False,
        "origen": "crm",
        "qr_token": str(uuid.uuid4()),
        "created_at": now,
        "updated_at": now
    }
    await db.orders.insert_one(order)
    await db.customers.update_one({"id": data.customer_id}, {"$inc": {"total_orders": 1}})
    await db.eventos_automation.insert_one({
        "id": str(uuid.uuid4()),
        "tipo": "ORDER_CREATED",
        "entity_id": order_id,
        "payload": {"order_number": order_number, "service_type": data.service_type},
        "created_at": now
    })
    await create_audit_log("ORDER_CREATED", "order", order_id, current_user["id"])
    await emit_realtime("notification", {
        "type": "order_created",
        "order_id": order_id,
        "status": "new",
        "order_number": order_number
    })
    
    # Send notifications if enabled
    if notify and NOTIFICATIONS_ENABLED and not SKIP_SERVER_NOTIFICATIONS:
        try:
            await notify_order_created(customer, order)
        except Exception as e:
            logger.error(f"Notification failed: {e}")
    
    return OrderResponse(**order)

@api_router.get("/orders", response_model=List[OrderResponse])
async def get_orders(
    status: Optional[str] = None,
    customer_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=100),
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        normalized = normalize_status(status)
        query["status"] = {"$in": [normalized, normalized.upper(), status]}
    if customer_id:
        query["customer_id"] = customer_id
    if date_from:
        query["created_at"] = {"$gte": date_from}
    if date_to:
        query.setdefault("created_at", {})["$lte"] = date_to
    
    skip = (page - 1) * page_size
    orders = await db.orders.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    return [OrderResponse(**o) for o in orders]

@api_router.get("/orders/{order_id}", response_model=OrderResponse)
async def get_order(order_id: str, current_user: dict = Depends(get_current_user)):
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    return OrderResponse(**order)

@api_router.get("/orders/{order_id}/qr")
async def get_order_qr(order_id: str, current_user: dict = Depends(get_current_user)):
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    qr_token = order.get("qr_token") or str(uuid.uuid4())
    if not order.get("qr_token"):
        await db.orders.update_one({"id": order_id}, {"$set": {"qr_token": qr_token}})
    return {"order_id": order_id, "order_number": order.get("order_number"), "qr_token": qr_token}

@api_router.get("/orders/{order_id}/qr.svg")
async def get_order_qr_svg(order_id: str, current_user: dict = Depends(get_current_user)):
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        order = await db.orders.find_one({"order_number": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    qr_token = order.get("qr_token") or str(uuid.uuid4())
    if not order.get("qr_token"):
        await db.orders.update_one({"id": order.get("id")}, {"$set": {"qr_token": qr_token}})

    customer = None
    if order.get("customer_id"):
        customer = await db.customers.find_one({"id": order.get("customer_id")}, {"_id": 0})

    payload = build_qr_payload({"id": order.get("id"), "order_number": order.get("order_number"), "qr_token": qr_token})
    ticket_svg = build_ticket_svg(order, customer, payload)
    display_id = build_display_order_number(order)
    filename = f"ticket-{display_id}.svg"
    return StreamingResponse(io.BytesIO(ticket_svg), media_type="image/svg+xml", headers={"Content-Disposition": f"attachment; filename={filename}"})

@api_router.get("/orders/qr/export")
async def export_qr_batch(
    start_date: str = Query(..., description="Start date YYYY-MM-DD"),
    end_date: str = Query(..., description="End date YYYY-MM-DD"),
    status: Optional[str] = None,
    service_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"pickup_date": {"$gte": start_date, "$lte": end_date}}
    if status:
        query["status"] = status
    if service_type:
        query["service_type"] = service_type
    export_limit = 500
    orders = await db.orders.find(query, {"_id": 0}).sort("pickup_date", 1).limit(export_limit + 1).to_list(export_limit + 1)
    if len(orders) > export_limit:
        raise HTTPException(status_code=400, detail=f"Export limit exceeded. Please narrow your filter to {export_limit} orders or fewer.")

    customer_ids = {order.get("customer_id") for order in orders if order.get("customer_id")}
    customer_map = {}
    if customer_ids:
        customers = await db.customers.find({"id": {"$in": list(customer_ids)}}, {"_id": 0}).to_list(len(customer_ids))
        customer_map = {customer.get("id"): customer for customer in customers}

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for order in orders:
            qr_token = order.get("qr_token") or str(uuid.uuid4())
            if not order.get("qr_token"):
                await db.orders.update_one({"id": order["id"]}, {"$set": {"qr_token": qr_token}})
            payload = build_qr_payload({"id": order["id"], "order_number": order.get("order_number"), "qr_token": qr_token})
            customer = customer_map.get(order.get("customer_id"))
            ticket_svg = build_ticket_svg(order, customer, payload)
            display_id = build_display_order_number(order)
            filename = f"ticket-{display_id}.svg"
            zip_file.writestr(filename, ticket_svg)
    buffer.seek(0)
    filename = f"qr-export-{start_date}-to-{end_date}.zip"
    return StreamingResponse(buffer, media_type="application/zip", headers={"Content-Disposition": f"attachment; filename={filename}"})

@api_router.post("/orders/qr/resolve")
async def resolve_qr(data: QrResolveRequest, current_user: dict = Depends(get_current_user)):
    payload_data = {}
    if data.payload:
        payload_data = parse_qr_payload(data.payload)
    qr_token = data.qr_token or payload_data.get("qr_token")
    order_id = payload_data.get("order_id")
    order_number = payload_data.get("order_number")
    if not qr_token and not order_id and not order_number:
        raise HTTPException(status_code=400, detail="QR sin datos válidos")
    order = None
    if qr_token:
        order = await db.orders.find_one({"qr_token": qr_token}, {"_id": 0})
    if not order and order_id:
        order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order and order_number:
        order = await db.orders.find_one({"order_number": order_number}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Orden no encontrada")
    if qr_token and order.get("qr_token") and order.get("qr_token") != qr_token:
        raise HTTPException(status_code=400, detail="QR no coincide con la orden")
    customer_name = order.get("customer_name")
    customer = None
    if not customer_name and order.get("customer_id"):
        customer = await db.customers.find_one({"id": order.get("customer_id")}, {"_id": 0})
        customer_name = customer.get("name") if customer else None
    delivery_address = order.get("delivery_address") or order.get("pickup_address")
    address_parts = build_address_parts(delivery_address)
    response = {
        "order_id": order.get("id"),
        "order_number": order.get("order_number"),
        "service_type": order.get("service_type"),
        "customer_name": customer_name,
        "address": address_parts,
        "request_datetime": order.get("created_at"),
        "status": order.get("status"),
        "items": order.get("items") or order.get("services_included") or order.get("products") or [],
        "total_amount": order.get("total_amount"),
        "special_instructions": order.get("notes") or order.get("special_instructions"),
        "pickup_date": order.get("pickup_date"),
        "pickup_time_window": order.get("pickup_time_window"),
        "payment_status": order.get("payment_status")
    }
    return response

@api_router.put("/orders/{order_id}", response_model=OrderResponse)
async def update_order(order_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    if "actual_lbs" in data:
        order_snapshot = await db.orders.find_one({"id": order_id}, {"_id": 0})
        if order_snapshot:
            customer_snapshot = None
            if order_snapshot.get("customer_id"):
                customer_snapshot = await db.customers.find_one({"id": order_snapshot.get("customer_id")}, {"_id": 0})
            temp_order = {**order_snapshot, **data}
            total_amount = calculate_service_amount(temp_order, customer_snapshot)
            if total_amount is not None:
                data["total_amount"] = total_amount
    result = await db.orders.update_one({"id": order_id}, {"$set": data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Order not found")
    
    await create_audit_log("ORDER_UPDATED", "order", order_id, current_user["id"], {"changes": list(data.keys())})
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    return OrderResponse(**order)

def normalize_status(value: Optional[str]) -> str:
    if not value:
        return ""
    return value.strip().lower().replace(" ", "_")


def normalize_payment_method(value: Optional[str]) -> str:
    if not value:
        return ""
    normalized = value.strip().lower()
    mapping = {
        "efectivo": "cash",
        "cash": "cash",
        "tarjeta": "card",
        "card": "card",
        "credito": "card",
        "débito": "card",
        "debito": "card",
        "transferencia": "transfer",
        "transfer": "transfer",
        "transferencia_bancaria": "transfer",
        "otro": "other",
        "other": "other"
    }
    return mapping.get(normalized, normalized)


def is_active_member(order: dict, customer: Optional[dict]) -> bool:
    status_value = ""
    if order:
        status_value = order.get("membership_status") or ""
    if customer:
        status_value = customer.get("membership_status") or status_value
    status_normalized = normalize_spaces(status_value).lower() if status_value else ""
    if status_normalized in ["inactive", "cancelled", "canceled", "expired"]:
        return False
    if status_normalized in ["active", "current", "paid", "yes", "true"]:
        return True
    plan = None
    if order:
        plan = order.get("membership_plan")
    if customer and not plan:
        plan = customer.get("membership_plan")
    return bool(plan)


def calculate_service_amount(order: dict, customer: Optional[dict]) -> Optional[float]:
    service_type = normalize_status(order.get("service_type") or "pickup_delivery")
    lbs_value = order.get("actual_lbs")
    if lbs_value is None:
        return None
    try:
        lbs_value = float(lbs_value)
    except Exception:
        return None
    if lbs_value <= 0:
        return None

    if service_type == "wash_fold":
        billable_lbs = max(lbs_value, 10)
        amount = billable_lbs * 2.25
    else:
        rate = 2.50 if is_active_member(order, customer) else 2.75
        amount = max(lbs_value * rate, 40)

    return round(float(amount), 2)


def should_notify_order_status(order: dict, status_value: str) -> bool:
    """Determine if order status change should trigger notification"""
    status_normalized = normalize_status(status_value)
    if not status_normalized or status_normalized == "new":
        return False

    service_type = normalize_status(order.get("service_type") or "pickup_delivery")
    if service_type in ["wash_fold", "self_service"]:
        return status_normalized == "ready"

    return status_normalized in ["confirmed", "pickup_scheduled", "ready", "out_for_delivery", "delivered"]

@api_router.patch("/orders/{order_id}/status")
async def update_order_status(order_id: str, status: str, notify: bool = True, current_user: dict = Depends(get_current_user)):
    valid_statuses = ["new", "confirmed", "pickup_scheduled", "picked_up", "processing", "ready", "out_for_delivery", "delivered", "completed", "cancelled"]
    normalized_status = normalize_status(status)
    if normalized_status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {valid_statuses}")

    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    current_status = normalize_status(order.get("status"))
    service_type = normalize_spaces(order.get("service_type") or "pickup_delivery").lower().replace(" ", "_")
    if normalized_status == "completed":
        if service_type in ["wash_fold", "wash_fold_dropoff", "self_service"]:
            if current_status not in ["ready", "completed"]:
                raise HTTPException(status_code=400, detail="Wash & Fold must be ready before it can be completed")
        elif current_status not in ["delivered", "completed", "out_for_delivery"]:
            raise HTTPException(status_code=400, detail="Order must be delivered before it can be completed")

    await db.orders.update_one(
        {"id": order_id},
        {
            "$set": {
                "status": normalized_status,
                "estado_actual": normalized_status,
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "tiempos.ultimo_cambio_estado": datetime.now(timezone.utc).isoformat(),
                f"tiempos.fechas_estado.{normalized_status}": datetime.now(timezone.utc).isoformat()
            }
        }
    )

    await create_audit_log("ORDER_STATUS_CHANGED", "order", order_id, current_user["id"], {"new_status": normalized_status})
    await db.eventos_automation.insert_one({
        "id": str(uuid.uuid4()),
        "tipo": "ORDER_STATUS_CHANGED",
        "entity_id": order_id,
        "payload": {"status": normalized_status},
        "created_at": datetime.now(timezone.utc).isoformat()
    })

    if notify and NOTIFICATIONS_ENABLED and not SKIP_SERVER_NOTIFICATIONS and order.get("customer_id") and should_notify_order_status(order, normalized_status):
        customer = await db.customers.find_one({"id": order["customer_id"]}, {"_id": 0})
        if customer:
            order["status"] = normalized_status
            try:
                await notify_order_status_changed(customer, order, normalized_status)
            except Exception as e:
                logger.error(f"Notification failed: {e}")

    await emit_realtime("notification", {
        "type": "order_status",
        "order_id": order_id,
        "status": normalized_status
    })

    return {"message": f"Order status updated to {normalized_status}"}

@api_router.patch("/orders/{order_id}/payment-status")
async def update_order_payment_status(order_id: str, status: str, current_user: dict = Depends(get_current_user)):
    """Update payment status of an order"""
    valid_statuses = ["pending", "paid", "refunded", "failed"]
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid payment status. Must be one of: {valid_statuses}")
    
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    result = await db.orders.update_one(
        {"id": order_id},
        {
            "$set": {
                "payment_status": status,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    await create_audit_log("ORDER_PAYMENT_STATUS_CHANGED", "order", order_id, current_user["id"], {"payment_status": status})
    await emit_realtime("notification", {
        "type": "order_payment",
        "order_id": order_id,
        "status": status
    })
    return {"message": f"Payment status updated to {status}"}


@api_router.post("/orders/{order_id}/payment")
async def capture_order_payment(
    order_id: str,
    data: OrderPaymentUpdate,
    current_user: dict = Depends(require_role([ROLE_OPERATOR]))
):
    """Capture payment details for an order"""
    method = normalize_payment_method(data.payment_method)
    allowed = ["cash", "card", "transfer", "other"]
    if method not in allowed:
        raise HTTPException(status_code=400, detail=f"Invalid payment method. Must be one of: {allowed}")

    order = await db.orders.find_one({"$or": [{"id": order_id}, {"order_number": order_id}]}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    total_amount = order.get("total_amount")
    amount_received = data.amount_received

    if method == "cash" and amount_received is None:
        raise HTTPException(status_code=400, detail="Amount received is required for cash payments")

    if amount_received is None:
        amount_received = total_amount

    if total_amount is not None and amount_received is not None:
        try:
            total_amount = float(total_amount)
            amount_received = float(amount_received)
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid amount values")
        if amount_received < total_amount:
            raise HTTPException(status_code=400, detail="Amount received cannot be less than total")

    change_due = None
    if method == "cash" and total_amount is not None and amount_received is not None:
        change_due = round(amount_received - total_amount, 2)

    now = datetime.now(timezone.utc).isoformat()
    update_data = {
        "payment_status": "paid",
        "payment_method": method,
        "amount_paid": amount_received,
        "change_due": change_due,
        "paid_at": now,
        "updated_at": now
    }

    await db.orders.update_one({"id": order.get("id")}, {"$set": update_data})
    await create_audit_log("ORDER_PAYMENT_CAPTURED", "order", order.get("id"), current_user["id"], update_data)
    await emit_realtime("notification", {
        "type": "order_payment",
        "order_id": order.get("id"),
        "status": "paid",
        "method": method
    })

    return {"ok": True, "order_id": order.get("id"), **update_data}


@api_router.post("/orders/{order_id}/stripe-checkout", response_model=OrderStripeCheckoutResponse)
async def create_order_stripe_checkout(
    order_id: str,
    data: OrderStripeCheckoutRequest,
    request: Request,
    current_user: dict = Depends(require_role([ROLE_OPERATOR]))
):
    if not STRIPE_CHECKOUT_AVAILABLE:
        raise HTTPException(status_code=503, detail="Stripe integration not available")
    order = await db.orders.find_one({"$or": [{"id": order_id}, {"order_number": order_id}]}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    customer = None
    if order.get("customer_id"):
        customer = await db.customers.find_one({"id": order.get("customer_id")}, {"_id": 0})

    amount = calculate_service_amount(order, customer)
    if amount is None:
        raise HTTPException(status_code=400, detail="Actual lbs required to calculate payment")

    stripe_api_key = os.environ.get("STRIPE_API_KEY")
    if not stripe_api_key:
        raise HTTPException(status_code=500, detail="Payment configuration error")

    host_url = str(request.base_url).rstrip("/")
    webhook_url = f"{host_url}/api/webhook/stripe"
    success_url = f"{data.origin_url}/admin/operator?session_id={{CHECKOUT_SESSION_ID}}&order_id={order.get('id')}"
    cancel_url = f"{data.origin_url}/admin/operator?order_id={order.get('id')}&status=cancelled"

    stripe_checkout = StripeCheckout(api_key=stripe_api_key, webhook_url=webhook_url)
    session_request = CheckoutSessionRequest(
        amount=amount,
        currency="usd",
        success_url=success_url,
        cancel_url=cancel_url,
        metadata={
            "order_id": order.get("id"),
            "order_number": order.get("order_number") or "",
            "type": "service_order"
        }
    )
    session: CheckoutSessionResponse = await stripe_checkout.create_checkout_session(session_request)

    now = datetime.now(timezone.utc).isoformat()
    payment_doc = {
        "id": str(uuid.uuid4()),
        "session_id": session.session_id,
        "order_id": order.get("id"),
        "entity_type": "service_order",
        "amount": amount,
        "currency": "usd",
        "status": "initiated",
        "payment_status": "pending",
        "metadata": session_request.metadata,
        "created_at": now,
        "updated_at": now
    }
    await db.payment_transactions.insert_one(payment_doc)

    await db.orders.update_one(
        {"id": order.get("id")},
        {
            "$set": {
                "total_amount": amount,
                "payment_status": "pending",
                "payment_method": "card",
                "updated_at": now
            }
        }
    )

    return OrderStripeCheckoutResponse(session_id=session.session_id, url=session.url, amount=amount, currency="usd")


@api_router.get("/orders/stripe/status/{session_id}")
async def get_order_stripe_status(
    session_id: str,
    request: Request,
    current_user: dict = Depends(require_role([ROLE_OPERATOR]))
):
    if not STRIPE_CHECKOUT_AVAILABLE:
        raise HTTPException(status_code=503, detail="Stripe integration not available")
    stripe_api_key = os.environ.get("STRIPE_API_KEY")
    if not stripe_api_key:
        raise HTTPException(status_code=500, detail="Payment configuration error")

    host_url = str(request.base_url).rstrip("/")
    webhook_url = f"{host_url}/api/webhook/stripe"
    stripe_checkout = StripeCheckout(api_key=stripe_api_key, webhook_url=webhook_url)
    status: CheckoutStatusResponse = await stripe_checkout.get_checkout_status(session_id)

    transaction = await db.payment_transactions.find_one({"session_id": session_id}, {"_id": 0})
    if not transaction:
        raise HTTPException(status_code=404, detail="Payment session not found")

    update_fields = {
        "status": status.status,
        "payment_status": status.payment_status,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.payment_transactions.update_one({"session_id": session_id}, {"$set": update_fields})

    if status.payment_status == "paid" and transaction.get("payment_status") != "paid":
        order_id = transaction.get("order_id") or status.metadata.get("order_id")
        if order_id:
            await db.orders.update_one(
                {"id": order_id},
                {
                    "$set": {
                        "payment_status": "paid",
                        "payment_method": "card",
                        "amount_paid": transaction.get("amount"),
                        "change_due": 0,
                        "paid_at": datetime.now(timezone.utc).isoformat(),
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }
                }
            )

    return {
        "status": status.status,
        "payment_status": status.payment_status,
        "amount_total": status.amount_total,
        "currency": status.currency,
        "metadata": status.metadata
    }

async def notify_last_completed_order(current_user: dict = Depends(get_current_user)):
    require_admin(current_user)
    last_order = await db.orders.find({"status": "completed"}, {"_id": 0}).sort("updated_at", -1).limit(1).to_list(1)
    if not last_order:
        raise HTTPException(status_code=404, detail="No completed orders found")
    order = last_order[0]
    customer_id = order.get("customer_id")
    if not customer_id:
        raise HTTPException(status_code=400, detail="Order missing customer")
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    if NOTIFICATIONS_ENABLED and not SKIP_SERVER_NOTIFICATIONS:
        await notify_order_status_changed(customer, order, "completed")
    await create_audit_log("ORDER_COMPLETED_NOTIFICATION_SENT", "order", order["id"], current_user["id"])
    return {"ok": True, "order_id": order["id"], "order_number": order.get("order_number")}

# ==================== QUOTES ====================

async def generate_quote_number():
    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    count = await db.quotes.count_documents({"quote_number": {"$regex": f"^QT-{today}"}})
    return f"QT-{today}-{str(count + 1).zfill(4)}"

@api_router.post("/quotes", response_model=QuoteResponse)
async def create_quote(data: QuoteCreate, current_user: dict = Depends(get_current_user)):
    quote_id = str(uuid.uuid4())
    quote_number = await generate_quote_number()
    now = datetime.now(timezone.utc).isoformat()
    
    quote = {
        "id": quote_id,
        "quote_number": quote_number,
        "company_name": data.company_name,
        "contact_name": data.contact_name,
        "email": data.email.lower() if data.email else None,
        "phone": data.phone,
        "industry": data.industry,
        "estimated_lbs_per_week": data.estimated_lbs_per_week,
        "service_needs": data.service_needs,
        "notes": data.notes,
        "status": "new",
        "assigned_to": None,
        "follow_up_date": None,
        "created_at": now,
        "updated_at": now
    }
    await db.quotes.insert_one(quote)
    await create_audit_log("QUOTE_CREATED", "quote", quote_id, current_user["id"])
    return QuoteResponse(**quote)

@api_router.get("/quotes", response_model=List[QuoteResponse])
async def get_quotes(
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query["status"] = status
    
    quotes = await db.quotes.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [QuoteResponse(**q) for q in quotes]

@api_router.get("/quotes/{quote_id}", response_model=QuoteResponse)
async def get_quote(quote_id: str, current_user: dict = Depends(get_current_user)):
    quote = await db.quotes.find_one({"id": quote_id}, {"_id": 0})
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")
    return QuoteResponse(**quote)

@api_router.put("/quotes/{quote_id}", response_model=QuoteResponse)
async def update_quote(quote_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.quotes.update_one({"id": quote_id}, {"$set": data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Quote not found")
    
    await create_audit_log("QUOTE_UPDATED", "quote", quote_id, current_user["id"])
    quote = await db.quotes.find_one({"id": quote_id}, {"_id": 0})
    return QuoteResponse(**quote)


@api_router.post("/quotes/{quote_id}/convert-to-lead", response_model=LeadResponse)
async def convert_quote_to_lead(quote_id: str, current_user: dict = Depends(get_current_user)):
    quote = await db.quotes.find_one({"id": quote_id}, {"_id": 0})
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")
    if quote.get("converted_lead_id"):
        raise HTTPException(status_code=400, detail="Quote already converted")

    lead_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    lead_name = quote.get("contact_name") or quote.get("company_name") or "B2B Lead"
    lead_notes = "\n".join([
        note for note in [quote.get("service_needs"), quote.get("notes")] if note
    ])

    lead = {
        "id": lead_id,
        "name": lead_name,
        "email": normalize_email(quote.get("email")) if quote.get("email") else None,
        "phone": normalize_phone(quote.get("phone")) if quote.get("phone") else None,
        "source": "b2b_quote",
        "interest_type": "B2B Quote",
        "notes": lead_notes or None,
        "status": "new",
        "converted_to_customer_id": None,
        "created_at": now,
        "updated_at": now
    }
    await db.leads.insert_one(lead)
    await db.quotes.update_one(
        {"id": quote_id},
        {"$set": {"status": "won", "converted_lead_id": lead_id, "updated_at": now}}
    )
    await create_audit_log("QUOTE_CONVERTED_TO_LEAD", "quote", quote_id, current_user["id"], {"lead_id": lead_id})
    return LeadResponse(**lead)

# ==================== LEADS ====================

@api_router.post("/leads", response_model=LeadResponse)
async def create_lead(data: LeadCreate, current_user: dict = Depends(get_current_user)):
    lead_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    lead = {
        "id": lead_id,
        "name": data.name,
        "email": data.email.lower() if data.email else None,
        "phone": data.phone,
        "source": data.source,
        "interest_type": data.interest_type,
        "notes": data.notes,
        "status": "new",
        "converted_to_customer_id": None,
        "created_at": now,
        "updated_at": now
    }
    await db.leads.insert_one(lead)
    await create_audit_log("LEAD_CREATED", "lead", lead_id, current_user["id"])
    return LeadResponse(**lead)

@api_router.get("/leads", response_model=List[LeadResponse])
async def get_leads(
    status: Optional[str] = None,
    source: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=100),
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query["status"] = status
    if source:
        query["source"] = source
    
    skip = (page - 1) * page_size
    leads = await db.leads.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    return [LeadResponse(**l) for l in leads]

@api_router.get("/leads/{lead_id}", response_model=LeadResponse)
async def get_lead(lead_id: str, current_user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    return LeadResponse(**lead)

@api_router.put("/leads/{lead_id}", response_model=LeadResponse)
async def update_lead(lead_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.leads.update_one({"id": lead_id}, {"$set": data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    await create_audit_log("LEAD_UPDATED", "lead", lead_id, current_user["id"])
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    return LeadResponse(**lead)

@api_router.post("/leads/{lead_id}/convert", response_model=CustomerResponse)
async def convert_lead_to_customer(lead_id: str, current_user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    if lead["status"] == "converted":
        raise HTTPException(status_code=400, detail="Lead already converted")
    
    customer_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    customer = {
        "id": customer_id,
        "name": lead["name"],
        "email": lead["email"],
        "phone": lead["phone"],
        "address": None,
        "preferred_contact": "email",
        "notes": lead.get("notes"),
        "status": "active",
        "total_orders": 0,
        "created_at": now,
        "updated_at": now
    }
    await db.customers.insert_one(customer)
    await db.leads.update_one(
        {"id": lead_id},
        {"$set": {"status": "converted", "converted_to_customer_id": customer_id, "updated_at": now}}
    )
    await create_audit_log("LEAD_CONVERTED", "lead", lead_id, current_user["id"], {"customer_id": customer_id})
    return CustomerResponse(**customer)

# ==================== SUPPORT TICKETS ====================

async def generate_ticket_number():
    count = await db.tickets.count_documents({})
    return f"TKT-{str(count + 1).zfill(5)}"

def determine_priority(subject: str, description: str) -> str:
    text = (subject + " " + description).lower()
    high_keywords = ["refund", "damaged", "missing", "complaint", "urgent", "broken", "lost"]
    if any(kw in text for kw in high_keywords):
        return "high"
    medium_keywords = ["issue", "problem", "error", "wrong"]
    if any(kw in text for kw in medium_keywords):
        return "medium"
    return "low"

@api_router.post("/tickets", response_model=TicketResponse)
async def create_ticket(data: TicketCreate, current_user: dict = Depends(get_current_user)):
    customer_name = None
    if data.customer_id:
        customer = await db.customers.find_one({"id": data.customer_id}, {"_id": 0})
        customer_name = customer["name"] if customer else None
    
    ticket_id = str(uuid.uuid4())
    ticket_number = await generate_ticket_number()
    priority = determine_priority(data.subject, data.description)
    now = datetime.now(timezone.utc).isoformat()
    
    ticket = {
        "id": ticket_id,
        "ticket_number": ticket_number,
        "customer_id": data.customer_id,
        "customer_name": customer_name,
        "subject": data.subject,
        "description": data.description,
        "category": data.category,
        "priority": priority,
        "status": "open",
        "assigned_to": None,
        "resolution": None,
        "created_at": now,
        "updated_at": now
    }
    await db.tickets.insert_one(ticket)
    await create_audit_log("TICKET_CREATED", "ticket", ticket_id, current_user["id"])
    return TicketResponse(**ticket)

@api_router.get("/tickets", response_model=List[TicketResponse])
async def get_tickets(
    status: Optional[str] = None,
    priority: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query["status"] = status
    if priority:
        query["priority"] = priority
    
    tickets = await db.tickets.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [TicketResponse(**t) for t in tickets]

@api_router.get("/tickets/{ticket_id}", response_model=TicketResponse)
async def get_ticket(ticket_id: str, current_user: dict = Depends(get_current_user)):
    ticket = await db.tickets.find_one({"id": ticket_id}, {"_id": 0})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    return TicketResponse(**ticket)

@api_router.put("/tickets/{ticket_id}", response_model=TicketResponse)
async def update_ticket(ticket_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.tickets.update_one({"id": ticket_id}, {"$set": data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    await create_audit_log("TICKET_UPDATED", "ticket", ticket_id, current_user["id"])
    ticket = await db.tickets.find_one({"id": ticket_id}, {"_id": 0})
    return TicketResponse(**ticket)

# ==================== SERVICES ====================

@api_router.post("/services", response_model=ServiceResponse)
async def create_service(data: ServiceCreate, current_user: dict = Depends(get_current_user)):
    service_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    service = {
        "id": service_id,
        "name": data.name,
        "category": data.category,
        "description": data.description,
        "price": data.price,
        "price_unit": data.price_unit,
        "is_active": data.is_active,
        "sort_order": data.sort_order or 0,
        "created_at": now,
        "updated_at": now
    }
    await db.services.insert_one(service)
    await create_audit_log("SERVICE_CREATED", "service", service_id, current_user["id"])
    return ServiceResponse(**service)

@api_router.get("/services", response_model=List[ServiceResponse])
async def get_services(
    active_only: bool = True,
    search: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if active_only:
        query["is_active"] = True
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"category": {"$regex": search, "$options": "i"}}
        ]
    services = await db.services.find(query, {"_id": 0}).sort([("sort_order", 1), ("created_at", -1)]).to_list(1000)
    return [ServiceResponse(**s) for s in services]

@api_router.get("/services/{service_id}", response_model=ServiceResponse)
async def get_service(service_id: str, current_user: dict = Depends(get_current_user)):
    service = await db.services.find_one({"id": service_id}, {"_id": 0})
    if not service:
        raise HTTPException(status_code=404, detail="Service not found")
    return ServiceResponse(**service)

@api_router.put("/services/{service_id}", response_model=ServiceResponse)
async def update_service(service_id: str, data: ServiceCreate, current_user: dict = Depends(get_current_user)):
    update_data = data.model_dump(exclude_unset=True)
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["sort_order"] = update_data.get("sort_order", 0) or 0
    result = await db.services.update_one({"id": service_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Service not found")
    await create_audit_log("SERVICE_UPDATED", "service", service_id, current_user["id"])
    service = await db.services.find_one({"id": service_id}, {"_id": 0})
    return ServiceResponse(**service)

@api_router.delete("/services/{service_id}")
async def delete_service(service_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.services.delete_one({"id": service_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Service not found")
    await create_audit_log("SERVICE_DELETED", "service", service_id, current_user["id"])
    return {"message": "Service deleted"}

@api_router.get("/public/services", response_model=List[ServiceResponse])
async def get_public_services(active_only: bool = True):
    query = {}
    if active_only:
        query["is_active"] = True
    services = await db.services.find(query, {"_id": 0}).sort([("sort_order", 1), ("created_at", -1)]).to_list(1000)
    return [ServiceResponse(**s) for s in services]

@api_router.get("/services/membership-section", response_model=MembershipSectionResponse)
async def get_membership_section(current_user: dict = Depends(get_current_user)):
    section = await db.membership_section.find_one({"id": "default"}, {"_id": 0})
    if not section:
        section = {
            "id": "default",
            "heading": "Flexible Plans for Every Home",
            "subheading": None,
            "special_title": "🎉 New Member Special",
            "special_text": "$10 OFF your first month on any membership. Ask when you call or text.",
            "cta_title": "Need help choosing?",
            "cta_text": "Just call, text, or email us at (805) 836-8872 and we'll recommend the perfect plan based on your weekly laundry.",
            "cta_button_label": "👉 BECOME A MEMBER",
            "cta_button_url": "/membership",
            "contact_phone": "(805) 836-8872",
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.membership_section.insert_one(section)
    return MembershipSectionResponse(**section)

@api_router.put("/services/membership-section", response_model=MembershipSectionResponse)
async def update_membership_section(data: MembershipSectionUpdate, current_user: dict = Depends(get_current_user)):
    now = datetime.now(timezone.utc).isoformat()
    update_data = data.model_dump(exclude_unset=True)
    update_data["updated_at"] = now
    await db.membership_section.update_one(
        {"id": "default"},
        {"$set": update_data, "$setOnInsert": {"id": "default", "created_at": now}},
        upsert=True
    )
    section = await db.membership_section.find_one({"id": "default"}, {"_id": 0})
    await create_audit_log("MEMBERSHIP_SECTION_UPDATED", "membership_section", "default", current_user["id"])
    return MembershipSectionResponse(**section)

@api_router.post("/services/membership-plans", response_model=MembershipPlanResponse)
async def create_membership_plan(data: MembershipPlanCreate, current_user: dict = Depends(get_current_user)):
    plan_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    plan = {
        "id": plan_id,
        "name": data.name,
        "price": data.price,
        "image_url": data.image_url,
        "features": data.features,
        "is_popular": data.is_popular,
        "is_active": data.is_active,
        "sort_order": data.sort_order or 0,
        "created_at": now,
        "updated_at": now
    }
    await db.membership_plans.insert_one(plan)
    await create_audit_log("MEMBERSHIP_PLAN_CREATED", "membership_plan", plan_id, current_user["id"])
    return MembershipPlanResponse(**plan)

@api_router.get("/services/membership-plans", response_model=List[MembershipPlanResponse])
async def get_membership_plans(
    active_only: bool = True,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if active_only:
        query["is_active"] = True
    plans = await db.membership_plans.find(query, {"_id": 0}).sort([("sort_order", 1), ("created_at", -1)]).to_list(200)
    if len(plans) == 0:
        now = datetime.now(timezone.utc).isoformat()
        seed = [
            {
                "id": str(uuid.uuid4()),
                "name": "MOST POPULAR",
                "price": "$139 / month",
                "image_url": "https://images.squarespace-cdn.com/content/v1/696c559a4b2b9b1b0febf8d7/4a2815a1-54c1-45fb-8320-244dce8b83c8/MOST+POPULAR.png",
                "features": ["Up to 60 lb/ month", "Basic Preferences saved (folding notes)", "Best value for most families"],
                "is_popular": True,
                "is_active": True,
                "sort_order": 1,
                "created_at": now,
                "updated_at": now
            },
            {
                "id": str(uuid.uuid4()),
                "name": "FAMILY PLUS",
                "price": "$199 / month",
                "image_url": "https://images.squarespace-cdn.com/content/v1/696c559a4b2b9b1b0febf8d7/f262a5b8-0043-4977-9d32-d6b343be3e70/FAMILY+PLUS.png",
                "features": ["Up to 90 lb/ month", "Priority scheduling", "Great for larger households or rentals"],
                "is_popular": False,
                "is_active": True,
                "sort_order": 2,
                "created_at": now,
                "updated_at": now
            },
            {
                "id": str(uuid.uuid4()),
                "name": "ELITE CONCIERGE",
                "price": "$299 / month",
                "image_url": "https://images.squarespace-cdn.com/content/v1/696c559a4b2b9b1b0febf8d7/13a4c501-7792-4f72-bf5c-072f95b5f995/ELITE+CONCIERGE.png",
                "features": ["Up to 120 lb/ month", "Priority turnaround (when possible)", "Premium packaging", "Saved preferences", "1 emergency pickup included"],
                "is_popular": False,
                "is_active": True,
                "sort_order": 3,
                "created_at": now,
                "updated_at": now
            }
        ]
        await db.membership_plans.insert_many(seed)
        plans = seed
    return [MembershipPlanResponse(**p) for p in plans]

@api_router.put("/services/membership-plans/{plan_id}", response_model=MembershipPlanResponse)
async def update_membership_plan(plan_id: str, data: MembershipPlanCreate, current_user: dict = Depends(get_current_user)):
    update_data = data.model_dump(exclude_unset=True)
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["sort_order"] = update_data.get("sort_order", 0) or 0
    result = await db.membership_plans.update_one({"id": plan_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Plan not found")
    await create_audit_log("MEMBERSHIP_PLAN_UPDATED", "membership_plan", plan_id, current_user["id"])
    plan = await db.membership_plans.find_one({"id": plan_id}, {"_id": 0})
    return MembershipPlanResponse(**plan)

@api_router.delete("/services/membership-plans/{plan_id}")
async def delete_membership_plan(plan_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.membership_plans.delete_one({"id": plan_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Plan not found")
    await create_audit_log("MEMBERSHIP_PLAN_DELETED", "membership_plan", plan_id, current_user["id"])
    return {"message": "Plan deleted"}

@api_router.get("/memberships/section", response_model=MembershipSectionResponse)
async def get_membership_section_admin(current_user: dict = Depends(get_current_user)):
    section = await db.membership_section.find_one({"id": "default"}, {"_id": 0})
    if not section:
        section = {
            "id": "default",
            "heading": "Flexible Plans for Every Home",
            "subheading": None,
            "special_title": "🎉 New Member Special",
            "special_text": "$10 OFF your first month on any membership. Ask when you call or text.",
            "cta_title": "Need help choosing?",
            "cta_text": "Just call, text, or email us at (805) 836-8872 and we'll recommend the perfect plan based on your weekly laundry.",
            "cta_button_label": "👉 BECOME A MEMBER",
            "cta_button_url": "/membership",
            "contact_phone": "(805) 836-8872",
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.membership_section.insert_one(section)
    return MembershipSectionResponse(**section)

@api_router.put("/memberships/section", response_model=MembershipSectionResponse)
async def update_membership_section_admin(data: MembershipSectionUpdate, current_user: dict = Depends(get_current_user)):
    now = datetime.now(timezone.utc).isoformat()
    update_data = data.model_dump(exclude_unset=True)
    update_data["updated_at"] = now
    await db.membership_section.update_one(
        {"id": "default"},
        {"$set": update_data, "$setOnInsert": {"id": "default", "created_at": now}},
        upsert=True
    )
    section = await db.membership_section.find_one({"id": "default"}, {"_id": 0})
    await create_audit_log("MEMBERSHIP_SECTION_UPDATED", "membership_section", "default", current_user["id"])
    return MembershipSectionResponse(**section)

@api_router.post("/memberships/plans", response_model=MembershipPlanResponse)
async def create_membership_plan_admin(data: MembershipPlanCreate, current_user: dict = Depends(get_current_user)):
    plan_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    plan = {
        "id": plan_id,
        "name": data.name,
        "price": data.price,
        "image_url": data.image_url,
        "features": data.features,
        "is_popular": data.is_popular,
        "is_active": data.is_active,
        "sort_order": data.sort_order or 0,
        "created_at": now,
        "updated_at": now
    }
    await db.membership_plans.insert_one(plan)
    await create_audit_log("MEMBERSHIP_PLAN_CREATED", "membership_plan", plan_id, current_user["id"])
    return MembershipPlanResponse(**plan)

@api_router.get("/memberships/plans", response_model=List[MembershipPlanResponse])
async def get_membership_plans_admin(
    active_only: bool = True,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if active_only:
        query["is_active"] = True
    plans = await db.membership_plans.find(query, {"_id": 0}).sort([("sort_order", 1), ("created_at", -1)]).to_list(200)
    if len(plans) == 0:
        now = datetime.now(timezone.utc).isoformat()
        seed = [
            {
                "id": str(uuid.uuid4()),
                "name": "MOST POPULAR",
                "price": "$139 / month",
                "image_url": "https://images.squarespace-cdn.com/content/v1/696c559a4b2b9b1b0febf8d7/4a2815a1-54c1-45fb-8320-244dce8b83c8/MOST+POPULAR.png",
                "features": ["Up to 60 lb/ month", "Basic Preferences saved (folding notes)", "Best value for most families"],
                "is_popular": True,
                "is_active": True,
                "sort_order": 1,
                "created_at": now,
                "updated_at": now
            },
            {
                "id": str(uuid.uuid4()),
                "name": "FAMILY PLUS",
                "price": "$199 / month",
                "image_url": "https://images.squarespace-cdn.com/content/v1/696c559a4b2b9b1b0febf8d7/f262a5b8-0043-4977-9d32-d6b343be3e70/FAMILY+PLUS.png",
                "features": ["Up to 90 lb/ month", "Priority scheduling", "Great for larger households or rentals"],
                "is_popular": False,
                "is_active": True,
                "sort_order": 2,
                "created_at": now,
                "updated_at": now
            },
            {
                "id": str(uuid.uuid4()),
                "name": "ELITE CONCIERGE",
                "price": "$299 / month",
                "image_url": "https://images.squarespace-cdn.com/content/v1/696c559a4b2b9b1b0febf8d7/13a4c501-7792-4f72-bf5c-072f95b5f995/ELITE+CONCIERGE.png",
                "features": ["Up to 120 lb/ month", "Priority turnaround (when possible)", "Premium packaging", "Saved preferences", "1 emergency pickup included"],
                "is_popular": False,
                "is_active": True,
                "sort_order": 3,
                "created_at": now,
                "updated_at": now
            }
        ]
        await db.membership_plans.insert_many(seed)
        plans = seed
    return [MembershipPlanResponse(**p) for p in plans]

@api_router.put("/memberships/plans/{plan_id}", response_model=MembershipPlanResponse)
async def update_membership_plan_admin(plan_id: str, data: MembershipPlanCreate, current_user: dict = Depends(get_current_user)):
    update_data = data.model_dump(exclude_unset=True)
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["sort_order"] = update_data.get("sort_order", 0) or 0
    result = await db.membership_plans.update_one({"id": plan_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Plan not found")
    await create_audit_log("MEMBERSHIP_PLAN_UPDATED", "membership_plan", plan_id, current_user["id"])
    plan = await db.membership_plans.find_one({"id": plan_id}, {"_id": 0})
    return MembershipPlanResponse(**plan)

@api_router.delete("/memberships/plans/{plan_id}")
async def delete_membership_plan_admin(plan_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.membership_plans.delete_one({"id": plan_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Plan not found")
    await create_audit_log("MEMBERSHIP_PLAN_DELETED", "membership_plan", plan_id, current_user["id"])
    return {"message": "Plan deleted"}

@api_router.get("/memberships/signups", response_model=List[MembershipSignupResponse])
async def get_membership_signups(
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    require_admin(current_user)
    query = {}
    if status:
        query["status"] = status
    signups = await db.membership_signups.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Transform signups to match response model
    result = []
    for s in signups:
        try:
            # Handle both old and new signup formats
            signup_data = {
                "id": s.get("id", ""),
                "first_name": s.get("first_name", ""),
                "last_name": s.get("last_name", ""),
                "email": s.get("email", s.get("customer_email", "")),
                "phone": s.get("phone", s.get("customer_phone", "")),
                "contact_method": s.get("contact_method", ""),
                "address_line1": s.get("address_line1", ""),
                "address_line2": s.get("address_line2"),
                "city": s.get("city", ""),
                "state": s.get("state", ""),
                "zip_code": s.get("zip_code", ""),
                "membership_plan": s.get("membership_plan", s.get("plan_name", "")),
                "plan_name": s.get("plan_name"),
                "plan_id": s.get("plan_id"),
                "laundry_frequency": s.get("laundry_frequency", ""),
                "estimated_lbs": s.get("estimated_lbs", 0) or 0,
                "amount": s.get("amount"),
                "payment_status": s.get("payment_status"),
                "status": s.get("status", "pending"),
                "customer_id": s.get("customer_id"),
                "customer_name": s.get("customer_name"),
                "customer_email": s.get("customer_email"),
                "customer_phone": s.get("customer_phone"),
                "stripe_session_id": s.get("stripe_session_id"),
                "preferences": s.get("preferences"),
                "created_at": s.get("created_at", ""),
                "updated_at": s.get("updated_at", "")
            }
            result.append(MembershipSignupResponse(**signup_data))
        except Exception as e:
            logger.error(f"Error processing signup {s.get('id')}: {e}")
            continue
    return result

@api_router.put("/memberships/signups/{signup_id}", response_model=MembershipSignupResponse)
async def update_membership_signup(signup_id: str, data: MembershipSignupUpdate, current_user: dict = Depends(get_current_user)):
    require_admin(current_user)
    update_data = data.model_dump(exclude_unset=True)
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.membership_signups.update_one({"id": signup_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Signup not found")
    await create_audit_log("MEMBERSHIP_SIGNUP_UPDATED", "membership_signup", signup_id, current_user["id"])
    signup = await db.membership_signups.find_one({"id": signup_id}, {"_id": 0})
    return MembershipSignupResponse(**signup)

@api_router.post("/memberships/signups/{signup_id}/convert", response_model=CustomerResponse)
async def convert_membership_signup(signup_id: str, current_user: dict = Depends(get_current_user)):
    require_admin(current_user)
    signup = await db.membership_signups.find_one({"id": signup_id}, {"_id": 0})
    if not signup:
        raise HTTPException(status_code=404, detail="Signup not found")
    now = datetime.now(timezone.utc).isoformat()
    customer = await db.customers.find_one({"email": signup["email"]}, {"_id": 0})
    if customer:
        update_data = {
            "membership_plan": signup["membership_plan"],
            "membership_status": "active",
            "membership_start_date": now,
            "updated_at": now
        }
        await db.customers.update_one({"id": customer["id"]}, {"$set": update_data})
        customer = await db.customers.find_one({"id": customer["id"]}, {"_id": 0})
    else:
        customer_id = str(uuid.uuid4())
        customer = {
            "id": customer_id,
            "name": f"{signup['first_name']} {signup['last_name']}",
            "email": signup["email"].lower(),
            "phone": signup["phone"],
            "address": f"{signup['address_line1']}{', ' + signup['address_line2'] if signup.get('address_line2') else ''}, {signup['city']}, {signup['state']} {signup['zip_code']}",
            "preferred_contact": signup["contact_method"],
            "notes": None,
            "status": "active",
            "total_orders": 0,
            "membership_plan": signup["membership_plan"],
            "membership_status": "active",
            "membership_start_date": now,
            "created_at": now,
            "updated_at": now
        }
        await db.customers.insert_one(customer)
        await create_audit_log("CUSTOMER_CREATED", "customer", customer_id, current_user["id"])

    preferences = signup.get("preferences")
    if preferences:
        existing_pref = await db.preferences.find({"customer_id": customer["id"]}).sort("version", -1).limit(1).to_list(1)
        version = (existing_pref[0]["version"] + 1) if existing_pref else 1
        pref_id = str(uuid.uuid4())
        normalized = normalize_preference_payload(PreferenceCreate(customer_id=customer["id"], **preferences))
        pref_doc = {
            "id": pref_id,
            "customer_id": customer["id"],
            **normalized,
            "version": version,
            "created_at": now,
            "updated_at": now
        }
        await db.preferences.insert_one(pref_doc)
        await db.customers.update_one({"id": customer["id"]}, {"$set": {"preferences_id": pref_id, "updated_at": now}})
    await db.membership_signups.update_one(
        {"id": signup_id},
        {"$set": {"status": "converted", "customer_id": customer["id"], "updated_at": now}}
    )
    await create_audit_log("MEMBERSHIP_SIGNUP_CONVERTED", "membership_signup", signup_id, current_user["id"], {"customer_id": customer["id"]})
    return CustomerResponse(**customer)

@api_router.get("/memberships/customers", response_model=List[CustomerResponse])
async def get_membership_customers(
    search: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    require_admin(current_user)
    query: Dict[str, Any] = {"membership_plan": {"$ne": None}}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}}
        ]
    customers = await db.customers.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [CustomerResponse(**c) for c in customers]

@api_router.put("/memberships/customers/{customer_id}", response_model=CustomerResponse)
async def update_membership_customer(customer_id: str, data: MembershipCustomerUpdate, current_user: dict = Depends(get_current_user)):
    require_admin(current_user)
    update_data = data.model_dump(exclude_unset=True)
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.customers.update_one({"id": customer_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    await create_audit_log("CUSTOMER_MEMBERSHIP_UPDATED", "customer", customer_id, current_user["id"])
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    return CustomerResponse(**customer)

# ==================== AI ASSISTANT ENDPOINTS ====================

@api_router.get("/ai/briefing")
async def get_daily_briefing(current_user: dict = Depends(get_current_user)):
    """Get AI-generated daily briefing for the current user"""
    if not AI_ASSISTANT_ENABLED:
        raise HTTPException(status_code=503, detail="AI Assistant not available")
    
    try:
        briefing = await generate_daily_briefing(
            db, 
            current_user.get("role", "operator"),
            current_user.get("name", "User")
        )
        return briefing
    except Exception as e:
        logger.error(f"Error generating briefing: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/ai/suggestions")
async def get_ai_suggestions(current_user: dict = Depends(get_current_user)):
    """Get AI-powered action suggestions"""
    if not AI_ASSISTANT_ENABLED:
        raise HTTPException(status_code=503, detail="AI Assistant not available")
    
    try:
        suggestions = await ai_suggest_actions(db, "general")
        return {"suggestions": suggestions}
    except Exception as e:
        logger.error(f"Error getting suggestions: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/ai/chat")
async def ai_chat(data: AdminAIRequest, current_user: dict = Depends(get_current_user)):
    """Chat with AI Business Assistant"""
    if not AI_ASSISTANT_ENABLED:
        raise HTTPException(status_code=503, detail="AI Assistant not available")
    
    try:
        result = await ai_analyze_business(
            db,
            data.message,
            current_user.get("role", "operator")
        )
        
        # Check if response contains action instructions
        response_text = result.get("response", "")
        actions = []
        results = []
        
        # Parse and execute actions if requested
        if data.execute and "```json" in response_text:
            import re
            json_matches = re.findall(r'```json\s*(\{.*?\})\s*```', response_text, re.DOTALL)
            for json_str in json_matches:
                try:
                    action_data = json.loads(json_str)
                    if "action" in action_data:
                        action_result = await execute_ai_action(action_data, current_user)
                        results.append(action_result)
                        actions.append(action_data)
                except:
                    pass
        
        return {
            "reply": response_text,
            "actions": actions,
            "results": results,
            "generated_at": result.get("generated_at")
        }
    except Exception as e:
        logger.error(f"Error in AI chat: {e}")
        raise HTTPException(status_code=500, detail=str(e))


CRITICAL_AI_ACTION_TYPES = {
    "register_payment",
    "update_user_role",
    "update_system_setting",
    "update_membership_customer",
    "update_store_payment_status"
}


def is_critical_ai_action(action_type: Optional[str], payload: Optional[dict]) -> bool:
    action_type = normalize_status(action_type or "").lower()
    payload = payload or {}
    if action_type in CRITICAL_AI_ACTION_TYPES:
        return True
    if action_type == "update_order_status":
        status_value = normalize_status(payload.get("status") or "")
        return status_value in {"CANCELLED", "COMPLETED"}
    if action_type == "update_store_order_status":
        status_value = normalize_status(payload.get("status") or "")
        return status_value in {"CANCELLED", "REFUNDED"}
    return False


async def get_or_create_ai_session(session_id: Optional[str], current_user: dict) -> dict:
    now = datetime.now(timezone.utc).isoformat()
    target_session_id = session_id or str(uuid.uuid4())
    session = await db.ai_operator_sessions.find_one({"session_id": target_session_id}, {"_id": 0})
    if session:
        return session

    session = {
        "session_id": target_session_id,
        "user_id": current_user.get("id"),
        "user_role": current_user.get("role"),
        "messages": [],
        "created_at": now,
        "updated_at": now
    }
    await db.ai_operator_sessions.insert_one(session)
    return session


async def save_ai_session_messages(session_id: str, messages: List[dict]):
    await db.ai_operator_sessions.update_one(
        {"session_id": session_id},
        {
            "$set": {
                "messages": messages[-80:],
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        },
        upsert=True
    )


async def build_global_operations_context() -> dict:
    orders = await db.orders.find({}, {"_id": 0, "id": 1, "order_number": 1, "status": 1, "customer_name": 1, "total_amount": 1, "payment_status": 1, "service_type": 1, "updated_at": 1}).sort("updated_at", -1).limit(30).to_list(30)
    store_orders = await db.store_orders.find({}, {"_id": 0, "id": 1, "order_number": 1, "status": 1, "payment_status": 1, "customer_name": 1, "total": 1, "updated_at": 1}).sort("updated_at", -1).limit(20).to_list(20)
    tickets = await db.tickets.find({}, {"_id": 0, "id": 1, "status": 1, "priority": 1, "created_at": 1}).sort("created_at", -1).limit(20).to_list(20)
    quotes = await db.quotes.find({}, {"_id": 0, "id": 1, "quote_number": 1, "status": 1, "company_name": 1, "updated_at": 1}).sort("updated_at", -1).limit(20).to_list(20)
    leads = await db.leads.find({}, {"_id": 0, "id": 1, "name": 1, "status": 1, "updated_at": 1}).sort("updated_at", -1).limit(20).to_list(20)
    signups = await db.membership_signups.find({}, {"_id": 0, "id": 1, "status": 1, "membership_plan": 1, "contact_name": 1, "updated_at": 1}).sort("updated_at", -1).limit(20).to_list(20)
    users = await db.users.find({}, {"_id": 0, "id": 1, "email": 1, "role": 1, "active": 1, "updated_at": 1}).sort("updated_at", -1).limit(20).to_list(20)

    stats = {
        "orders_total": await db.orders.count_documents({}),
        "orders_processing": await db.orders.count_documents({"status": {"$in": ["processing", "PROCESSING"]}}),
        "orders_ready": await db.orders.count_documents({"status": {"$in": ["ready", "READY"]}}),
        "store_orders_pending_payment": await db.store_orders.count_documents({"payment_status": {"$in": ["pending", "pending_payment", "unpaid"]}}),
        "tickets_open": await db.tickets.count_documents({"status": {"$in": ["open", "OPEN"]}}),
        "quotes_open": await db.quotes.count_documents({"status": {"$nin": ["closed", "rejected", "converted"]}}),
        "leads_open": await db.leads.count_documents({"status": {"$nin": ["won", "lost", "closed"]}})
    }

    return {
        "stats": stats,
        "orders": orders,
        "store_orders": store_orders,
        "tickets": tickets,
        "quotes": quotes,
        "leads": leads,
        "signups": signups,
        "users": users,
        "generated_at": datetime.now(timezone.utc).isoformat()
    }


async def try_direct_charge_answer(message: str) -> Optional[dict]:
    text = normalize_spaces(message or "")
    if not text:
        return None

    lower = text.lower()
    trigger_phrases = [
        "cuánto le cobro a", "cuanto le cobro a", "how much should i charge", "how much to charge"
    ]
    if not any(phrase in lower for phrase in trigger_phrases):
        return None

    name_guess = None
    match_es = re.search(r"(?:cuánto le cobro a|cuanto le cobro a)\s+([\w\sÁÉÍÓÚÜÑáéíóúüñ'-]+)", text, flags=re.IGNORECASE)
    match_en = re.search(r"(?:how much should i charge|how much to charge)\s+([\w\s'-]+)", text, flags=re.IGNORECASE)
    if match_es:
        name_guess = normalize_spaces(match_es.group(1))
    elif match_en:
        name_guess = normalize_spaces(match_en.group(1))

    if not name_guess:
        return None

    customer = await db.customers.find_one(
        {"name": {"$regex": name_guess, "$options": "i"}},
        {"_id": 0, "id": 1, "name": 1}
    )
    if not customer:
        return {
            "reply": f"No encontré un cliente llamado {name_guess} en el sistema. ¿Quieres que busque por teléfono, correo o número de orden para calcular el cobro exacto?",
            "actions": []
        }

    orders = await db.orders.find(
        {"customer_id": customer.get("id")},
        {"_id": 0, "order_number": 1, "total_amount": 1, "payment_status": 1, "status": 1, "updated_at": 1}
    ).sort("updated_at", -1).limit(5).to_list(5)

    if not orders:
        return {
            "reply": f"No encontré órdenes activas para {customer.get('name')}. ¿Quieres que cree una nueva orden o revise historial más antiguo?",
            "actions": []
        }

    latest = orders[0]
    due_amount = 0.0
    for order in orders:
        payment_status = normalize_status(order.get("payment_status") or "")
        if payment_status not in {"PAID", "SETTLED"}:
            try:
                due_amount += float(order.get("total_amount") or 0)
            except Exception:
                continue

    due_text = f"${due_amount:.2f}" if due_amount > 0 else "$0.00"
    latest_total = latest.get("total_amount")
    latest_total_text = f"${float(latest_total):.2f}" if latest_total is not None else "N/A"
    latest_status = latest.get("status") or "N/A"
    latest_order_number = latest.get("order_number") or "N/A"

    reply = (
        f"Para {customer.get('name')}, el saldo pendiente estimado es {due_text}. "
        f"La orden más reciente es {latest_order_number} con total {latest_total_text} y estado {latest_status}. "
        f"¿Quieres que registre pago ahora o te detalle cada orden pendiente?"
    )
    return {"reply": reply, "actions": []}


async def execute_jarvis_action(action_type: str, action_payload: dict, current_user: dict, base_url: str) -> dict:
    action_type = normalize_status(action_type or "").lower()
    action_payload = action_payload or {}

    if action_type == "update_order_status":
        order_id = action_payload.get("order_id")
        status = action_payload.get("status")
        if not order_id or not status:
            return {"type": action_type, "ok": False, "error": "Missing order_id or status"}
        result = await update_order_status(order_id, status, True, current_user)
        return {"type": action_type, "ok": True, "order_id": order_id, "status": status, "result": result}

    if action_type == "update_order_lbs":
        order_id = action_payload.get("order_id")
        if not order_id:
            return {"type": action_type, "ok": False, "error": "Missing order_id"}
        update_data = {}
        if "estimated_lbs" in action_payload:
            update_data["estimated_lbs"] = action_payload.get("estimated_lbs")
        if "actual_lbs" in action_payload:
            update_data["actual_lbs"] = action_payload.get("actual_lbs")
        if not update_data:
            return {"type": action_type, "ok": False, "error": "No lbs provided"}
        updated = await update_order(order_id, update_data, current_user)
        return {"type": action_type, "ok": True, "order_id": order_id, "updated": updated}

    if action_type == "register_payment":
        order_id = action_payload.get("order_id")
        method = action_payload.get("payment_method")
        amount_received = action_payload.get("amount_received")
        if not order_id or not method:
            return {"type": action_type, "ok": False, "error": "Missing payment data"}
        payment_request = OrderPaymentUpdate(payment_method=method, amount_received=amount_received)
        updated = await capture_order_payment(order_id, payment_request, current_user)
        return {"type": action_type, "ok": True, "order_id": order_id, "updated": updated}

    if action_type == "print_ticket":
        order_id = action_payload.get("order_id")
        if not order_id:
            return {"type": action_type, "ok": False, "error": "Missing order_id"}
        ticket_url = f"{base_url}/api/orders/{order_id}/qr.svg"
        return {"type": action_type, "ok": True, "order_id": order_id, "ticket_url": ticket_url}

    if action_type == "send_notification":
        order_id = action_payload.get("order_id")
        message = action_payload.get("message")
        channel = action_payload.get("channel")
        if not order_id:
            return {"type": action_type, "ok": False, "error": "Missing order_id"}
        order = await db.orders.find_one({"$or": [{"id": order_id}, {"order_number": order_id}]}, {"_id": 0})
        if not order:
            return {"type": action_type, "ok": False, "error": "Order not found"}
        customer = await db.customers.find_one({"id": order.get("customer_id")}, {"_id": 0}) if order.get("customer_id") else None
        if not customer:
            return {"type": action_type, "ok": False, "error": "Customer not found"}

        if message:
            preferred = normalize_preferred_contact(channel or customer.get("preferred_contact") or order.get("preferred_contact"))
            if preferred == "email":
                ok = await send_email(customer.get("email"), "Order update", message)
            elif preferred == "call":
                language = detect_language(customer, customer.get("phone"))
                ok = await send_voice_call(customer.get("phone"), message, language)
            elif preferred == "whatsapp":
                ok = await send_whatsapp(customer.get("phone"), message)
            else:
                ok = await send_sms(customer.get("phone"), message)
        else:
            ok = await send_preferred_notification(customer, order, "status_changed", order.get("status"))
        return {"type": action_type, "ok": ok, "order_id": order_id}

    if action_type == "update_ticket_status":
        ticket_id = action_payload.get("ticket_id")
        status = action_payload.get("status")
        if not ticket_id or not status:
            return {"type": action_type, "ok": False, "error": "Missing ticket_id or status"}
        result = await db.tickets.update_one({"id": ticket_id}, {"$set": {"status": status, "updated_at": datetime.now(timezone.utc).isoformat()}})
        return {"type": action_type, "ok": result.modified_count > 0, "ticket_id": ticket_id, "status": status}

    if action_type == "update_quote_status":
        quote_id = action_payload.get("quote_id")
        status = action_payload.get("status")
        if not quote_id or not status:
            return {"type": action_type, "ok": False, "error": "Missing quote_id or status"}
        result = await db.quotes.update_one({"id": quote_id}, {"$set": {"status": status, "updated_at": datetime.now(timezone.utc).isoformat()}})
        return {"type": action_type, "ok": result.modified_count > 0, "quote_id": quote_id, "status": status}

    if action_type == "update_lead_status":
        lead_id = action_payload.get("lead_id")
        status = action_payload.get("status")
        if not lead_id or not status:
            return {"type": action_type, "ok": False, "error": "Missing lead_id or status"}
        result = await db.leads.update_one({"id": lead_id}, {"$set": {"status": status, "updated_at": datetime.now(timezone.utc).isoformat()}})
        return {"type": action_type, "ok": result.modified_count > 0, "lead_id": lead_id, "status": status}

    if action_type == "update_signup_status":
        signup_id = action_payload.get("signup_id")
        status = action_payload.get("status")
        if not signup_id or not status:
            return {"type": action_type, "ok": False, "error": "Missing signup_id or status"}
        result = await db.membership_signups.update_one({"id": signup_id}, {"$set": {"status": status, "updated_at": datetime.now(timezone.utc).isoformat()}})
        return {"type": action_type, "ok": result.modified_count > 0, "signup_id": signup_id, "status": status}

    if action_type == "update_membership_customer":
        customer_id = action_payload.get("customer_id")
        update_fields = {}
        for key in ["membership_plan", "membership_status", "membership_start_date"]:
            if action_payload.get(key) is not None:
                update_fields[key] = action_payload.get(key)
        if not customer_id or not update_fields:
            return {"type": action_type, "ok": False, "error": "Missing customer_id or membership fields"}
        update_fields["updated_at"] = datetime.now(timezone.utc).isoformat()
        result = await db.customers.update_one({"id": customer_id}, {"$set": update_fields})
        return {"type": action_type, "ok": result.modified_count > 0, "customer_id": customer_id, "updated_fields": list(update_fields.keys())}

    if action_type == "update_store_order_status":
        order_id = action_payload.get("order_id")
        status = action_payload.get("status")
        if not order_id or not status:
            return {"type": action_type, "ok": False, "error": "Missing order_id or status"}
        result = await db.store_orders.update_one({"$or": [{"id": order_id}, {"order_number": order_id}]}, {"$set": {"status": status, "updated_at": datetime.now(timezone.utc).isoformat()}})
        return {"type": action_type, "ok": result.modified_count > 0, "order_id": order_id, "status": status}

    if action_type == "update_store_payment_status":
        order_id = action_payload.get("order_id")
        payment_status = action_payload.get("payment_status")
        if not order_id or not payment_status:
            return {"type": action_type, "ok": False, "error": "Missing order_id or payment_status"}
        result = await db.store_orders.update_one(
            {"$or": [{"id": order_id}, {"order_number": order_id}]},
            {"$set": {"payment_status": payment_status, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
        return {"type": action_type, "ok": result.modified_count > 0, "order_id": order_id, "payment_status": payment_status}

    if action_type == "update_user_role":
        user_id = action_payload.get("user_id")
        role = action_payload.get("role")
        if not user_id or role not in VALID_ROLES:
            return {"type": action_type, "ok": False, "error": "Missing user_id or invalid role"}
        result = await db.users.update_one({"id": user_id}, {"$set": {"role": role, "updated_at": datetime.now(timezone.utc).isoformat()}})
        return {"type": action_type, "ok": result.modified_count > 0, "user_id": user_id, "role": role}

    if action_type == "update_system_setting":
        setting_key = normalize_spaces(action_payload.get("key") or "")
        value = action_payload.get("value")
        if not setting_key:
            return {"type": action_type, "ok": False, "error": "Missing setting key"}
        await db.system_settings.update_one(
            {"key": setting_key},
            {"$set": {"value": value, "updated_at": datetime.now(timezone.utc).isoformat(), "updated_by": current_user.get("id")}},
            upsert=True
        )
        return {"type": action_type, "ok": True, "key": setting_key}

    return {"type": action_type, "ok": False, "error": "Unknown action"}

@api_router.post("/ai/operations")
async def ai_operations(
    data: AdminAIRequest,
    request: Request,
    current_user: dict = Depends(require_role([ROLE_OPERATOR, ROLE_ADMIN]))
):
    """Jarvis-style global operations assistant with persistent memory and guarded execution."""
    if not AI_ASSISTANT_ENABLED:
        raise HTTPException(status_code=503, detail="AI Assistant not available")
    session = await get_or_create_ai_session(data.session_id, current_user)
    session_id = session.get("session_id")
    history = session.get("messages", [])[-20:]
    context_bundle = await build_global_operations_context()

    direct_answer = await try_direct_charge_answer(data.message)
    if direct_answer:
        now = datetime.now(timezone.utc).isoformat()
        history.append({"role": "user", "content": data.message, "created_at": now})
        history.append({"role": "assistant", "content": direct_answer.get("reply"), "created_at": now, "actions": []})
        await save_ai_session_messages(session_id, history)
        await db.ai_command_logs.insert_one({
            "id": str(uuid.uuid4()),
            "session_id": session_id,
            "user_id": current_user.get("id"),
            "user_role": current_user.get("role"),
            "message": data.message,
            "reply": direct_answer.get("reply"),
            "actions": [],
            "critical_actions": [],
            "requires_confirmation": False,
            "confirm_token": None,
            "executed": False,
            "results": [],
            "confidence": 1,
            "created_at": now,
            "source": "direct_charge_answer"
        })
        return {
            "session_id": session_id,
            "reply": direct_answer.get("reply"),
            "actions": [],
            "critical_actions": [],
            "requires_confirmation": False,
            "confirm_token": None,
            "results": [],
            "global_context": context_bundle.get("stats", {}),
            "generated_at": now
        }

    history_lines = "\n".join([
        f"{item.get('role', 'user')}: {item.get('content', '')}" for item in history if item.get("content")
    ]) or "No prior session history"

    system_prompt = (
        "You are JARVIS, the formal-professional omnichannel operations AI for Ventura Fresh Laundry. "
        "You must answer any user query clearly, but only suggest executable actions that exist in the system. "
        "You have global context across orders, store, tickets, quotes, leads, users, settings and finance indicators. "
        "Return ONLY valid JSON with keys: reply (string), actions (array), confidence (number 0-1). "
        "Each action must include: type, payload, reason."
        "Allowed action types: update_order_status, update_order_lbs, register_payment, print_ticket, send_notification, "
        "update_ticket_status, update_quote_status, update_lead_status, update_signup_status, update_membership_customer, "
        "update_store_order_status, update_store_payment_status, update_user_role, update_system_setting. "
        "Be concise, strategic, and accurate."
    )

    prompt = (
        f"{system_prompt}\n\n"
        f"SESSION HISTORY:\n{history_lines}\n\n"
        f"GLOBAL CONTEXT JSON:\n{json.dumps(context_bundle, ensure_ascii=False)}\n\n"
        f"CLIENT CONTEXT JSON:\n{json.dumps(data.context or {}, ensure_ascii=False)}\n\n"
        f"USER MESSAGE:\n{data.message}\n\n"
        f"JSON:"
    )

    ai_error = None
    try:
        raw = call_ollama(prompt)
        try:
            payload = extract_json_payload(raw)
        except Exception:
            payload = {"reply": raw, "actions": [], "confidence": 0.5}
    except Exception as exc:
        ai_error = str(exc)
        context_locale = normalize_spaces((data.context or {}).get("locale") if isinstance(data.context, dict) else "")
        is_spanish = (context_locale or "").lower().startswith("es") or any(ch in (data.message or "") for ch in ["¿", "á", "é", "í", "ó", "ú", "ñ"])
        fallback_reply = (
            "Estoy temporalmente en modo de alta demanda de IA. Puedo seguir ejecutando comandos directos del sistema "
            "si me das la acción exacta (por ejemplo: actualizar orden X a READY, registrar pago, o cerrar ticket Y)."
            if is_spanish else
            "I'm temporarily in high-demand AI mode. I can still execute direct system commands if you provide an explicit action "
            "(for example: update order X to READY, register payment, or close ticket Y)."
        )
        payload = {"reply": fallback_reply, "actions": [], "confidence": 0.2}

    reply = normalize_spaces(payload.get("reply") or "Done")
    actions = payload.get("actions", []) if isinstance(payload.get("actions", []), list) else []
    confidence = payload.get("confidence", 0.5)

    host = request.headers.get("x-forwarded-host") or request.headers.get("host")
    proto = request.headers.get("x-forwarded-proto") or "https"
    base_url = f"{proto}://{host}" if host else str(request.base_url).rstrip("/")

    critical_actions = []
    for action in actions:
        action_type = action.get("type") or action.get("action")
        action_payload = action.get("payload") or action.get("params") or {}
        if is_critical_ai_action(action_type, action_payload):
            critical_actions.append({"type": action_type, "payload": action_payload})

    requires_confirmation = len(critical_actions) > 0 and data.execute
    confirmation_token = data.confirm_token
    results = []

    if requires_confirmation:
        pending = await db.ai_pending_actions.find_one({"session_id": session_id, "status": "pending"}, {"_id": 0})
        if pending and pending.get("token") == confirmation_token:
            await db.ai_pending_actions.update_one(
                {"session_id": session_id, "token": confirmation_token},
                {"$set": {"status": "confirmed", "confirmed_at": datetime.now(timezone.utc).isoformat(), "confirmed_by": current_user.get("id")}}
            )
            requires_confirmation = False
        elif not confirmation_token:
            confirmation_token = str(uuid.uuid4())
            await db.ai_pending_actions.insert_one({
                "id": str(uuid.uuid4()),
                "session_id": session_id,
                "token": confirmation_token,
                "status": "pending",
                "actions": actions,
                "critical_actions": critical_actions,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": current_user.get("id")
            })

    if data.execute and not requires_confirmation:
        for action in actions:
            action_type = action.get("type") or action.get("action")
            action_payload = action.get("payload") or action.get("params") or {}
            try:
                result = await execute_jarvis_action(action_type, action_payload, current_user, base_url)
                results.append(result)
            except Exception as exc:
                results.append({"type": action_type, "ok": False, "error": str(exc)})

    now = datetime.now(timezone.utc).isoformat()
    history.append({"role": "user", "content": data.message, "created_at": now})
    history.append({"role": "assistant", "content": reply, "created_at": now, "actions": actions, "results": results})
    await save_ai_session_messages(session_id, history)

    await db.ai_command_logs.insert_one({
        "id": str(uuid.uuid4()),
        "session_id": session_id,
        "user_id": current_user.get("id"),
        "user_role": current_user.get("role"),
        "message": data.message,
        "reply": reply,
        "actions": actions,
        "critical_actions": critical_actions,
        "requires_confirmation": requires_confirmation,
        "confirm_token": confirmation_token if requires_confirmation else None,
        "executed": data.execute and not requires_confirmation,
        "results": results,
        "confidence": confidence,
        "ai_error": ai_error,
        "created_at": now
    })

    day_key = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    await db.ai_daily_summaries.update_one(
        {"day": day_key},
        {
            "$set": {"updated_at": now},
            "$inc": {
                "interactions_count": 1,
                "actions_proposed_count": len(actions),
                "actions_executed_count": len([r for r in results if r.get("ok")])
            },
            "$push": {
                "samples": {
                    "$each": [{
                        "session_id": session_id,
                        "user_id": current_user.get("id"),
                        "message": data.message,
                        "reply": reply,
                        "created_at": now
                    }],
                    "$slice": -50
                }
            }
        },
        upsert=True
    )

    response_reply = reply
    if requires_confirmation:
        response_reply = f"{reply}\n\nCritical actions detected. Please confirm execution to proceed."

    return {
        "session_id": session_id,
        "reply": response_reply,
        "actions": actions,
        "critical_actions": critical_actions,
        "requires_confirmation": requires_confirmation,
        "confirm_token": confirmation_token if requires_confirmation else None,
        "results": results,
        "global_context": context_bundle.get("stats", {}),
        "generated_at": now
    }


@api_router.get("/ai/operations/session/{session_id}")
async def get_ai_operations_session(session_id: str, current_user: dict = Depends(require_role([ROLE_OPERATOR, ROLE_ADMIN]))):
    session = await db.ai_operator_sessions.find_one({"session_id": session_id}, {"_id": 0})
    if not session:
        return {"session_id": session_id, "messages": []}

    if session.get("user_id") and session.get("user_id") != current_user.get("id") and current_user.get("role") != ROLE_ADMIN:
        raise HTTPException(status_code=403, detail="Session not accessible")

    return {
        "session_id": session.get("session_id"),
        "messages": session.get("messages", [])[-60:],
        "updated_at": session.get("updated_at")
    }


async def execute_ai_action(action_data: dict, current_user: dict) -> dict:
    """Execute an action suggested by AI"""
    action_type = action_data.get("action")
    params = action_data.get("params", {})
    
    try:
        if action_type == "update_order_status":
            order_id = params.get("order_id")
            status = params.get("status")
            if order_id and status:
                await db.orders.update_one(
                    {"id": order_id},
                    {"$set": {"status": status, "updated_at": datetime.now(timezone.utc).isoformat()}}
                )
                await create_audit_log("ORDER_STATUS_CHANGED", "order", order_id, current_user["id"], {"status": status, "source": "ai"})
                return {"action": action_type, "ok": True, "message": f"Order updated to {status}"}

        elif action_type == "update_order_lbs":
            order_id = params.get("order_id")
            if order_id:
                update_data = {}
                if "estimated_lbs" in params:
                    update_data["estimated_lbs"] = params.get("estimated_lbs")
                if "actual_lbs" in params:
                    update_data["actual_lbs"] = params.get("actual_lbs")
                if update_data:
                    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
                    await db.orders.update_one({"id": order_id}, {"$set": update_data})
                    await create_audit_log("ORDER_UPDATED", "order", order_id, current_user["id"], {"changes": list(update_data.keys()), "source": "ai"})
                    return {"action": action_type, "ok": True, "message": "Order lbs updated"}

        elif action_type == "update_payment_status":
            order_id = params.get("order_id")
            status = params.get("status")
            if order_id and status:
                await db.orders.update_one(
                    {"id": order_id},
                    {"$set": {"payment_status": status, "updated_at": datetime.now(timezone.utc).isoformat()}}
                )
                await create_audit_log("ORDER_PAYMENT_UPDATED", "order", order_id, current_user["id"], {"payment_status": status, "source": "ai"})
                return {"action": action_type, "ok": True, "message": f"Payment status updated to {status}"}
        
        elif action_type == "update_quote_status":
            quote_id = params.get("quote_id")
            status = params.get("status")
            if quote_id and status:
                await db.quotes.update_one(
                    {"id": quote_id},
                    {"$set": {"status": status, "updated_at": datetime.now(timezone.utc).isoformat()}}
                )
                await create_audit_log("QUOTE_STATUS_CHANGED", "quote", quote_id, current_user["id"], {"status": status, "source": "ai"})
                return {"action": action_type, "ok": True, "message": f"Quote updated to {status}"}
        
        elif action_type == "update_lead_status":
            lead_id = params.get("lead_id")
            status = params.get("status")
            if lead_id and status:
                await db.leads.update_one(
                    {"id": lead_id},
                    {"$set": {"status": status, "updated_at": datetime.now(timezone.utc).isoformat()}}
                )
                await create_audit_log("LEAD_STATUS_CHANGED", "lead", lead_id, current_user["id"], {"status": status, "source": "ai"})
                return {"action": action_type, "ok": True, "message": f"Lead updated to {status}"}
        
        elif action_type == "update_ticket_status":
            ticket_id = params.get("ticket_id")
            status = params.get("status")
            if ticket_id and status:
                await db.tickets.update_one(
                    {"id": ticket_id},
                    {"$set": {"status": status, "updated_at": datetime.now(timezone.utc).isoformat()}}
                )
                await create_audit_log("TICKET_STATUS_CHANGED", "ticket", ticket_id, current_user["id"], {"status": status, "source": "ai"})
                return {"action": action_type, "ok": True, "message": f"Ticket updated to {status}"}
        
        return {"action": action_type, "ok": False, "error": "Unknown action type"}
    except Exception as e:
        return {"action": action_type, "ok": False, "error": str(e)}

@api_router.post("/admin/ai")
async def admin_ai(data: AdminAIRequest, current_user: dict = Depends(get_current_user)):
    require_admin(current_user)

    recent_orders = await db.orders.find({}, {"_id": 0}).sort("created_at", -1).limit(15).to_list(15)
    orders_summary = "\n".join([
        f"- {o.get('order_number') or o.get('id')} | id:{o.get('id')} | {o.get('customer_name') or '-'} | status:{o.get('status')} | est:{o.get('estimated_lbs')} | act:{o.get('actual_lbs')} | pickup:{o.get('pickup_date')}"
        for o in recent_orders
    ]) or "No recent orders"

    recent_quotes = await db.quotes.find({}, {"_id": 0}).sort("created_at", -1).limit(10).to_list(10)
    quotes_summary = "\n".join([
        f"- {q.get('quote_number') or q.get('id')} | id:{q.get('id')} | {q.get('company_name') or q.get('contact_name') or '-'} | status:{q.get('status')} | est_lbs:{q.get('estimated_lbs')}"
        for q in recent_quotes
    ]) or "No recent quotes"

    recent_leads = await db.leads.find({}, {"_id": 0}).sort("created_at", -1).limit(10).to_list(10)
    leads_summary = "\n".join([
        f"- {l.get('id')} | {l.get('name') or '-'} | status:{l.get('status')}"
        for l in recent_leads
    ]) or "No recent leads"

    context = f"RECENT ORDERS:\n{orders_summary}\n\nRECENT QUOTES:\n{quotes_summary}\n\nRECENT LEADS:\n{leads_summary}"

    system_prompt = (
        "You are a local admin assistant for Ventura Fresh Laundry CRM. "
        "Return ONLY valid JSON with keys: reply (string) and actions (array). "
        "Actions must be objects with type and payload. Allowed types: "
        "update_order_status, update_order_lbs, update_payment_status, update_ticket_status, update_quote_status, update_lead_status, "
        "update_membership_signup_status, update_customer_membership. "
        "For update_order_status payload: order_id, status (new|processing|ready|out_for_delivery|delivered|completed|cancelled). "
        "For update_order_lbs payload: order_id, estimated_lbs (number or null), actual_lbs (number or null). "
        "For update_payment_status payload: order_id, status (pending|paid|refunded|failed). "
        "For update_ticket_status payload: ticket_id, status. "
        "For update_quote_status payload: quote_id, status. "
        "For update_lead_status payload: lead_id, status. "
        "For update_membership_signup_status payload: signup_id, status. "
        "For update_customer_membership payload: customer_id, membership_plan, membership_status, membership_start_date. "
        "Use IDs from the CONTEXT. If no action is needed, return actions: []."
    )
    prompt = f"{system_prompt}\n\nCONTEXT:\n{context}\n\nUser: {data.message}\nJSON:"
    model_response = call_ollama(prompt)
    try:
        payload = extract_json_payload(model_response)
    except Exception:
        return {"reply": model_response, "actions": [], "results": []}
    reply = payload.get("reply", "")
    actions = payload.get("actions", []) if isinstance(payload.get("actions", []), list) else []
    results = []
    if data.execute:
        for action in actions:
            action_type = action.get("type")
            action_payload = action.get("payload", {})
            if action_type == "update_order_status":
                order_id = action_payload.get("order_id")
                status_value = action_payload.get("status")
                valid_statuses = ["new", "confirmed", "pickup_scheduled", "picked_up", "processing", "ready", "out_for_delivery", "delivered", "completed", "cancelled"]
                if not order_id or status_value not in valid_statuses:
                    results.append({"type": action_type, "ok": False, "error": "Invalid order status or id"})
                    continue
                now_iso = datetime.now(timezone.utc).isoformat()
                result = await db.orders.update_one(
                    {"id": order_id},
                    {
                        "$set": {
                            "status": status_value,
                            "estado_actual": status_value,
                            "updated_at": now_iso,
                            "tiempos.ultimo_cambio_estado": now_iso,
                            f"tiempos.fechas_estado.{status_value}": now_iso
                        }
                    }
                )
                if result.matched_count == 0:
                    results.append({"type": action_type, "ok": False, "error": "Order not found"})
                    continue
                await create_audit_log("ORDER_UPDATED", "order", order_id, current_user["id"], {"status": status_value, "source": "ai"})
                await db.eventos_automation.insert_one({
                    "id": str(uuid.uuid4()),
                    "tipo": "ORDER_STATUS_CHANGED",
                    "entity_id": order_id,
                    "payload": {"status": status_value, "source": "ai"},
                    "created_at": now_iso
                })
                if NOTIFICATIONS_ENABLED:
                    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
                    if order and order.get("customer_id") and should_notify_order_status(order, status_value):
                        customer = await db.customers.find_one({"id": order["customer_id"]}, {"_id": 0})
                        if customer:
                            try:
                                await notify_order_status_changed(customer, order, status_value)
                            except Exception as e:
                                logger.error(f"Notification failed: {e}")
                results.append({"type": action_type, "ok": True, "order_id": order_id, "status": status_value})
            elif action_type == "update_order_lbs":
                order_id = action_payload.get("order_id")
                if not order_id:
                    results.append({"type": action_type, "ok": False, "error": "Invalid order id"})
                    continue
                def to_float(value):
                    if value is None or value == "":
                        return None
                    try:
                        return float(value)
                    except Exception:
                        return None
                update_data = {}
                if "estimated_lbs" in action_payload:
                    update_data["estimated_lbs"] = to_float(action_payload.get("estimated_lbs"))
                if "actual_lbs" in action_payload:
                    update_data["actual_lbs"] = to_float(action_payload.get("actual_lbs"))
                if not update_data:
                    results.append({"type": action_type, "ok": False, "error": "No lbs provided"})
                    continue
                update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
                result = await db.orders.update_one({"id": order_id}, {"$set": update_data})
                if result.matched_count == 0:
                    results.append({"type": action_type, "ok": False, "error": "Order not found"})
                    continue
                await create_audit_log("ORDER_UPDATED", "order", order_id, current_user["id"], {"changes": list(update_data.keys()), "source": "ai"})
                results.append({"type": action_type, "ok": True, "order_id": order_id, "estimated_lbs": update_data.get("estimated_lbs"), "actual_lbs": update_data.get("actual_lbs")})
            elif action_type == "update_payment_status":
                order_id = action_payload.get("order_id")
                status_value = action_payload.get("status")
                valid_payment = ["pending", "paid", "refunded", "failed"]
                if not order_id or status_value not in valid_payment:
                    results.append({"type": action_type, "ok": False, "error": "Invalid payment status or id"})
                    continue
                result = await db.orders.update_one(
                    {"id": order_id},
                    {"$set": {"payment_status": status_value, "updated_at": datetime.now(timezone.utc).isoformat()}}
                )
                if result.matched_count == 0:
                    results.append({"type": action_type, "ok": False, "error": "Order not found"})
                    continue
                await create_audit_log("ORDER_PAYMENT_STATUS_CHANGED", "order", order_id, current_user["id"], {"status": status_value, "source": "ai"})
                results.append({"type": action_type, "ok": True, "order_id": order_id, "status": status_value})
            elif action_type == "update_ticket_status":
                ticket_id = action_payload.get("ticket_id")
                status_value = action_payload.get("status")
                if not ticket_id or not status_value:
                    results.append({"type": action_type, "ok": False, "error": "Invalid ticket status or id"})
                    continue
                result = await db.tickets.update_one({"id": ticket_id}, {"$set": {"status": status_value, "updated_at": datetime.now(timezone.utc).isoformat()}})
                if result.matched_count == 0:
                    results.append({"type": action_type, "ok": False, "error": "Ticket not found"})
                    continue
                await create_audit_log("TICKET_UPDATED", "ticket", ticket_id, current_user["id"], {"status": status_value, "source": "ai"})
                results.append({"type": action_type, "ok": True, "ticket_id": ticket_id, "status": status_value})
            elif action_type == "update_quote_status":
                quote_id = action_payload.get("quote_id")
                status_value = action_payload.get("status")
                if not quote_id or not status_value:
                    results.append({"type": action_type, "ok": False, "error": "Invalid quote status or id"})
                    continue
                result = await db.quotes.update_one({"id": quote_id}, {"$set": {"status": status_value, "updated_at": datetime.now(timezone.utc).isoformat()}})
                if result.matched_count == 0:
                    results.append({"type": action_type, "ok": False, "error": "Quote not found"})
                    continue
                await create_audit_log("QUOTE_UPDATED", "quote", quote_id, current_user["id"], {"status": status_value, "source": "ai"})
                results.append({"type": action_type, "ok": True, "quote_id": quote_id, "status": status_value})
            elif action_type == "update_lead_status":
                lead_id = action_payload.get("lead_id")
                status_value = action_payload.get("status")
                if not lead_id or not status_value:
                    results.append({"type": action_type, "ok": False, "error": "Invalid lead status or id"})
                    continue
                result = await db.leads.update_one({"id": lead_id}, {"$set": {"status": status_value, "updated_at": datetime.now(timezone.utc).isoformat()}})
                if result.matched_count == 0:
                    results.append({"type": action_type, "ok": False, "error": "Lead not found"})
                    continue
                await create_audit_log("LEAD_UPDATED", "lead", lead_id, current_user["id"], {"status": status_value, "source": "ai"})
                results.append({"type": action_type, "ok": True, "lead_id": lead_id, "status": status_value})
            elif action_type == "update_membership_signup_status":
                signup_id = action_payload.get("signup_id")
                status_value = action_payload.get("status")
                if not signup_id or not status_value:
                    results.append({"type": action_type, "ok": False, "error": "Invalid signup status or id"})
                    continue
                result = await db.membership_signups.update_one({"id": signup_id}, {"$set": {"status": status_value, "updated_at": datetime.now(timezone.utc).isoformat()}})
                if result.matched_count == 0:
                    results.append({"type": action_type, "ok": False, "error": "Signup not found"})
                    continue
                await create_audit_log("MEMBERSHIP_SIGNUP_UPDATED", "membership_signup", signup_id, current_user["id"], {"status": status_value, "source": "ai"})
                results.append({"type": action_type, "ok": True, "signup_id": signup_id, "status": status_value})
            elif action_type == "update_customer_membership":
                customer_id = action_payload.get("customer_id")
                if not customer_id:
                    results.append({"type": action_type, "ok": False, "error": "Invalid customer id"})
                    continue
                update_data = {k: v for k, v in action_payload.items() if k in ["membership_plan", "membership_status", "membership_start_date"] and v is not None}
                if not update_data:
                    results.append({"type": action_type, "ok": False, "error": "No membership fields provided"})
                    continue
                update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
                result = await db.customers.update_one({"id": customer_id}, {"$set": update_data})
                if result.matched_count == 0:
                    results.append({"type": action_type, "ok": False, "error": "Customer not found"})
                    continue
                await create_audit_log("CUSTOMER_MEMBERSHIP_UPDATED", "customer", customer_id, current_user["id"], {"source": "ai"})
                results.append({"type": action_type, "ok": True, "customer_id": customer_id})
            else:
                results.append({"type": action_type, "ok": False, "error": "Unsupported action"})
    return {"reply": reply, "actions": actions, "results": results}

@api_router.post("/admin/ai/insights")
async def admin_ai_insights(data: AdminAIInsightsRequest, current_user: dict = Depends(get_current_user)):
    require_admin(current_user)
    now = datetime.now(timezone.utc)
    today = now.strftime("%Y-%m-%d")
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()

    orders_by_status = {}
    for status_value in ["new", "processing", "ready", "out_for_delivery", "delivered", "completed", "cancelled"]:
        orders_by_status[status_value] = await db.orders.count_documents({"status": status_value})

    orders_today = await db.orders.count_documents({"created_at": {"$regex": f"^{today}"}})
    tickets_open = await db.tickets.count_documents({"status": {"$in": ["open", "in_progress"]}})
    quotes_new = await db.quotes.count_documents({"status": "new"})
    leads_new = await db.leads.count_documents({"status": "new"})
    signups_new = await db.membership_signups.count_documents({"status": "new"})

    pipeline = [
        {"$match": {"created_at": {"$gte": month_start}, "payment_status": "paid"}},
        {"$group": {"_id": None, "total": {"$sum": "$total_amount"}}}
    ]
    revenue_result = await db.orders.aggregate(pipeline).to_list(1)
    revenue = revenue_result[0]["total"] if revenue_result else 0

    latest_orders = await db.orders.find({}, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)

    snapshot = {
        "generated_at": now.isoformat(),
        "orders_today": orders_today,
        "orders_by_status": orders_by_status,
        "tickets_open": tickets_open,
        "quotes_new": quotes_new,
        "leads_new": leads_new,
        "membership_signups_new": signups_new,
        "revenue_this_month": revenue or 0,
        "latest_orders": [
            {
                "order_number": o.get("order_number"),
                "status": o.get("status"),
                "created_at": o.get("created_at"),
                "customer_name": o.get("customer_name")
            }
            for o in latest_orders
        ]
    }

    if data.type == "summary":
        prompt = (
            "Genera un resumen ejecutivo breve en español, en 4-6 líneas, "
            "con prioridades operativas usando este snapshot JSON:\n"
            f"{json.dumps(snapshot, ensure_ascii=False)}"
        )
    elif data.type == "risks":
        prompt = (
            "Analiza riesgos operativos y financieros en español con este snapshot JSON. "
            "Devuelve 5 bullets claros con riesgo y recomendación breve:\n"
            f"{json.dumps(snapshot, ensure_ascii=False)}"
        )
    elif data.type == "forecast":
        prompt = (
            "Genera una predicción corta en español sobre carga de trabajo y tendencias "
            "para la próxima semana usando este snapshot JSON. "
            "Incluye 3-5 bullets accionables:\n"
            f"{json.dumps(snapshot, ensure_ascii=False)}"
        )
    else:
        raise HTTPException(status_code=400, detail="Tipo de análisis inválido")

    reply = call_ollama(prompt)
    return {"reply": reply, "snapshot": snapshot}

@api_router.get("/finances/summary")
async def get_finances_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    require_admin(current_user)

    def parse_amount(value) -> float:
        try:
            return float(value)
        except (TypeError, ValueError):
            return 0.0

    def normalize_date(date_value: Optional[str], is_end: bool = False) -> Optional[str]:
        if not date_value:
            return None
        try:
            dt = datetime.fromisoformat(date_value)
        except ValueError:
            return None
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        if is_end:
            dt = dt.replace(hour=23, minute=59, second=59, microsecond=999999)
        else:
            dt = dt.replace(hour=0, minute=0, second=0, microsecond=0)
        return dt.isoformat()

    start_iso = normalize_date(start_date)
    end_iso = normalize_date(end_date, is_end=True)

    order_query: Dict[str, Any] = {}
    membership_query: Dict[str, Any] = {}
    if start_iso or end_iso:
        range_query: Dict[str, Any] = {}
        if start_iso:
            range_query["$gte"] = start_iso
        if end_iso:
            range_query["$lte"] = end_iso
        order_query["created_at"] = range_query
        membership_query["created_at"] = range_query

    orders = await db.orders.find(order_query, {"_id": 0}).to_list(5000)
    paid_orders = [o for o in orders if (o.get("payment_status") or "").lower() == "paid"]
    pending_orders = [o for o in orders if (o.get("payment_status") or "").lower() != "paid"]

    order_revenue = sum(parse_amount(o.get("total_amount")) for o in paid_orders)
    avg_order_value = order_revenue / len(paid_orders) if paid_orders else 0

    signups = await db.membership_signups.find(membership_query, {"_id": 0}).to_list(5000)
    paid_signups = [s for s in signups if (s.get("payment_status") or "").lower() == "paid"]
    membership_revenue = sum(parse_amount(s.get("amount")) for s in paid_signups)

    store_orders = await db.store_orders.find(order_query, {"_id": 0}).to_list(5000)
    paid_store_orders = [o for o in store_orders if (o.get("payment_status") or "").lower() == "paid"]
    pending_store_orders = [o for o in store_orders if (o.get("payment_status") or "").lower() != "paid"]
    store_revenue = sum(parse_amount(o.get("total")) for o in paid_store_orders)

    payment_methods = {}
    def add_payment(method, amount):
        key = (method or "unknown").lower()
        payment_methods.setdefault(key, {"count": 0, "amount": 0.0})
        payment_methods[key]["count"] += 1
        payment_methods[key]["amount"] += amount

    for order in paid_orders:
        add_payment(order.get("payment_method"), parse_amount(order.get("total_amount")))
    for order in paid_store_orders:
        add_payment(order.get("payment_method"), parse_amount(order.get("total")))

    total_revenue = order_revenue + membership_revenue + store_revenue

    return {
        "start_date": start_date,
        "end_date": end_date,
        "total_revenue": total_revenue,
        "order_revenue": order_revenue,
        "membership_revenue": membership_revenue,
        "store_revenue": store_revenue,
        "total_orders": len(orders),
        "paid_orders": len(paid_orders),
        "pending_orders": len(pending_orders),
        "store_orders": len(store_orders),
        "store_paid_orders": len(paid_store_orders),
        "store_pending_orders": len(pending_store_orders),
        "avg_order_value": avg_order_value,
        "total_memberships": len(paid_signups),
        "payment_methods": payment_methods
    }

@api_router.post("/ai/patrones/scan")
async def scan_patterns(data: PatternScanRequest, current_user: dict = Depends(get_current_user)):
    require_admin(current_user)
    await ensure_ai_indexes()
    now = datetime.now(timezone.utc)
    periodo_desde = data.periodo_desde or (now - timedelta(days=7)).isoformat()
    periodo_hasta = data.periodo_hasta or now.isoformat()
    filtros = data.filtros or {}
    base_match = {"created_at": {"$gte": periodo_desde, "$lte": periodo_hasta}}
    if "service_type" in filtros:
        base_match["service_type"] = {"$in": filtros["service_type"]}
    patterns = []
    processing_pipeline = [
        {"$match": base_match},
        {
            "$project": {
                "service_type": 1,
                "processing_ts": {"$dateFromString": {"dateString": "$tiempos.fechas_estado.processing"}},
                "ready_ts": {"$dateFromString": {"dateString": "$tiempos.fechas_estado.ready"}}
            }
        },
        {"$match": {"processing_ts": {"$ne": None}, "ready_ts": {"$ne": None}}},
        {
            "$project": {
                "service_type": 1,
                "duration_hours": {
                    "$divide": [{"$subtract": ["$ready_ts", "$processing_ts"]}, 1000 * 60 * 60]
                }
            }
        },
        {
            "$group": {
                "_id": "$service_type",
                "avg_hours": {"$avg": "$duration_hours"},
                "max_hours": {"$max": "$duration_hours"},
                "count": {"$sum": 1}
            }
        }
    ]
    processing_stats = await db.orders.aggregate(processing_pipeline).to_list(50)
    for stat in processing_stats:
        patterns.append({
            "id": str(uuid.uuid4()),
            "tipo": "cuello_botella",
            "detalle": {
                "estado": "processing",
                "service_type": stat["_id"],
                "avg_hours": stat["avg_hours"],
                "max_hours": stat["max_hours"]
            },
            "query_base": base_match,
            "periodo": {"desde": periodo_desde, "hasta": periodo_hasta},
            "fecha_deteccion": now.isoformat(),
            "impacto_estimado": {"ordenes_afectadas": stat["count"]}
        })
    errors_pipeline = [
        {"$match": base_match},
        {"$unwind": "$errores_validacion"},
        {
            "$group": {
                "_id": "$errores_validacion.codigo",
                "count": {"$sum": 1}
            }
        },
        {"$sort": {"count": -1}},
        {"$limit": 10}
    ]
    errors_stats = await db.orders.aggregate(errors_pipeline).to_list(10)
    for stat in errors_stats:
        patterns.append({
            "id": str(uuid.uuid4()),
            "tipo": "error_recurrente",
            "detalle": {"codigo": stat["_id"], "count": stat["count"]},
            "query_base": base_match,
            "periodo": {"desde": periodo_desde, "hasta": periodo_hasta},
            "fecha_deteccion": now.isoformat(),
            "impacto_estimado": {"ordenes_afectadas": stat["count"]}
        })
    stale_threshold = (now - timedelta(hours=48)).isoformat()
    stale_match = {**base_match, "status": {"$in": ["processing", "ready", "out_for_delivery"]}, "created_at": {"$lte": stale_threshold}}
    stale_pipeline = [
        {"$match": stale_match},
        {"$group": {"_id": "$status", "count": {"$sum": 1}}}
    ]
    stale_stats = await db.orders.aggregate(stale_pipeline).to_list(10)
    for stat in stale_stats:
        patterns.append({
            "id": str(uuid.uuid4()),
            "tipo": "desviacion",
            "detalle": {"estado": stat["_id"], "threshold_hours": 48},
            "query_base": stale_match,
            "periodo": {"desde": periodo_desde, "hasta": periodo_hasta},
            "fecha_deteccion": now.isoformat(),
            "impacto_estimado": {"ordenes_afectadas": stat["count"]}
        })
    if patterns:
        await db.patrones_detectados.insert_many(patterns)
    return {"ok": True, "patrones_creados": len(patterns)}

@api_router.post("/ai/propuestas/generar")
async def generate_proposals(data: ProposalGenerateRequest, current_user: dict = Depends(get_current_user)):
    require_admin(current_user)
    await ensure_ai_indexes()
    await get_or_seed_business_rules()
    query = {}
    if data.patrones_ids:
        query["id"] = {"$in": data.patrones_ids}
    patrones = await db.patrones_detectados.find(query, {"_id": 0}).sort("fecha_deteccion", -1).to_list(50)
    if not patrones:
        return {"ok": True, "propuestas_creadas": 0, "detalle": "Sin patrones"}
    prompt = (
        "Eres un asistente de optimización. Devuelve JSON con clave propuestas (array). "
        "Cada propuesta: tipo, descripcion, impacto_estimado, accion_sugerida, nivel_riesgo, datos_respaldo.\n"
        f"patrones={json.dumps(patrones, ensure_ascii=False)}"
    )
    raw = call_ollama(prompt)
    propuestas = []
    try:
        payload = extract_json_payload(raw)
        propuestas = payload.get("propuestas", [])
    except Exception:
        propuestas = []
    if not propuestas:
        for pattern in patrones[: data.max_propuestas or 10]:
            tipo = "ajuste_regla" if pattern["tipo"] == "cuello_botella" else "nueva_validacion"
            propuestas.append({
                "tipo": tipo,
                "descripcion": f"Propuesta automática basada en patrón {pattern['tipo']}",
                "impacto_estimado": pattern.get("impacto_estimado", {}),
                "accion_sugerida": {"type": tipo, "payload": {"pattern_id": pattern["id"]}},
                "nivel_riesgo": "medio",
                "datos_respaldo": {"pattern_id": pattern["id"]}
            })
    now = datetime.now(timezone.utc).isoformat()
    docs = []
    for propuesta in propuestas[: data.max_propuestas or 10]:
        docs.append({
            "id": str(uuid.uuid4()),
            "tipo": propuesta.get("tipo"),
            "descripcion": propuesta.get("descripcion", ""),
            "estado": "pendiente",
            "impacto_estimado": propuesta.get("impacto_estimado", {}),
            "accion_sugerida": propuesta.get("accion_sugerida", {}),
            "nivel_riesgo": propuesta.get("nivel_riesgo", "medio"),
            "datos_respaldo": propuesta.get("datos_respaldo", {}),
            "fecha_generacion": now,
            "fuente": "analizador_automatico_v1"
        })
    if docs:
        await db.propuestas_ia.insert_many(docs)
    return {"ok": True, "propuestas_creadas": len(docs)}

@api_router.get("/ai/propuestas")
async def list_proposals(estado: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    require_admin(current_user)
    await ensure_ai_indexes()
    query = {}
    if estado:
        query["estado"] = estado
    propuestas = await db.propuestas_ia.find(query, {"_id": 0}).sort("fecha_generacion", -1).to_list(200)
    return propuestas

@api_router.get("/ai/propuestas/{propuesta_id}")
async def get_proposal(propuesta_id: str, current_user: dict = Depends(get_current_user)):
    require_admin(current_user)
    propuesta = await db.propuestas_ia.find_one({"id": propuesta_id}, {"_id": 0})
    if not propuesta:
        raise HTTPException(status_code=404, detail="Propuesta no encontrada")
    return propuesta

@api_router.get("/ai/propuestas/{propuesta_id}/simulacion")
async def get_proposal_simulation(
    propuesta_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    service_type: Optional[str] = None,
    status: Optional[str] = None,
    real_before_days: Optional[int] = 14,
    real_after_days: Optional[int] = 7,
    current_user: dict = Depends(get_current_user)
):
    require_admin(current_user)
    propuesta = await db.propuestas_ia.find_one({"id": propuesta_id}, {"_id": 0})
    if not propuesta:
        raise HTTPException(status_code=404, detail="Propuesta no encontrada")
    now = datetime.now(timezone.utc)
    periodo_desde = start_date or (now - timedelta(days=7)).isoformat()
    periodo_hasta = end_date or now.isoformat()
    match = {"created_at": {"$gte": periodo_desde, "$lte": periodo_hasta}}
    inferred_service = propuesta.get("accion_sugerida", {}).get("payload", {}).get("service_type")
    service_type = service_type or inferred_service
    if service_type:
        match["service_type"] = service_type
    if status:
        match["status"] = status
    processing_pipeline = [
        {"$match": match},
        {
            "$project": {
                "processing_ts": {"$dateFromString": {"dateString": "$tiempos.fechas_estado.processing"}},
                "ready_ts": {"$dateFromString": {"dateString": "$tiempos.fechas_estado.ready"}}
            }
        },
        {"$match": {"processing_ts": {"$ne": None}, "ready_ts": {"$ne": None}}},
        {
            "$project": {
                "duration_hours": {
                    "$divide": [{"$subtract": ["$ready_ts", "$processing_ts"]}, 1000 * 60 * 60]
                }
            }
        },
        {"$group": {"_id": None, "avg_hours": {"$avg": "$duration_hours"}}}
    ]
    processing_stats = await db.orders.aggregate(processing_pipeline).to_list(1)
    avg_processing_hours = processing_stats[0]["avg_hours"] if processing_stats else 0
    errors_pipeline = [
        {"$match": match},
        {"$unwind": "$errores_validacion"},
        {"$group": {"_id": None, "count": {"$sum": 1}}}
    ]
    errors_stats = await db.orders.aggregate(errors_pipeline).to_list(1)
    errors_count = errors_stats[0]["count"] if errors_stats else 0
    orders_count = await db.orders.count_documents(match)
    before = {
        "ordenes": orders_count,
        "avg_processing_horas": round(avg_processing_hours, 2) if avg_processing_hours else 0,
        "errores_validacion": errors_count
    }
    after = dict(before)
    impacto = propuesta.get("impacto_estimado", {})
    tiempo_pct = impacto.get("tiempo_ahorrado_porcentaje")
    errores_pct = impacto.get("errores_reducidos_porcentaje")
    if isinstance(tiempo_pct, (int, float)):
        after["avg_processing_horas"] = round(before["avg_processing_horas"] * (1 - tiempo_pct / 100), 2)
    if isinstance(errores_pct, (int, float)):
        after["errores_validacion"] = max(0, round(before["errores_validacion"] * (1 - errores_pct / 100)))
    service_pipeline = [
        {"$match": match},
        {"$group": {"_id": "$service_type", "count": {"$sum": 1}}}
    ]
    service_stats = await db.orders.aggregate(service_pipeline).to_list(20)
    status_pipeline = [
        {"$match": match},
        {"$group": {"_id": "$status", "count": {"$sum": 1}}}
    ]
    status_stats = await db.orders.aggregate(status_pipeline).to_list(20)
    real_before_start = (now - timedelta(days=real_before_days or 14)).isoformat()
    real_before_end = (now - timedelta(days=real_after_days or 7)).isoformat()
    real_after_start = (now - timedelta(days=real_after_days or 7)).isoformat()
    real_after_end = now.isoformat()
    real_before_match = {"created_at": {"$gte": real_before_start, "$lte": real_before_end}}
    real_after_match = {"created_at": {"$gte": real_after_start, "$lte": real_after_end}}
    if service_type:
        real_before_match["service_type"] = service_type
        real_after_match["service_type"] = service_type
    if status:
        real_before_match["status"] = status
        real_after_match["status"] = status
    real_before_stats = await db.orders.aggregate([
        {"$match": real_before_match},
        {"$unwind": {"path": "$errores_validacion", "preserveNullAndEmptyArrays": True}},
        {
            "$group": {
                "_id": None,
                "errores": {"$sum": {"$cond": [{"$ifNull": ["$errores_validacion", False]}, 1, 0]}},
                "ordenes": {"$addToSet": "$id"}
            }
        }
    ]).to_list(1)
    real_after_stats = await db.orders.aggregate([
        {"$match": real_after_match},
        {"$unwind": {"path": "$errores_validacion", "preserveNullAndEmptyArrays": True}},
        {
            "$group": {
                "_id": None,
                "errores": {"$sum": {"$cond": [{"$ifNull": ["$errores_validacion", False]}, 1, 0]}},
                "ordenes": {"$addToSet": "$id"}
            }
        }
    ]).to_list(1)
    real_before = real_before_stats[0] if real_before_stats else {"errores": 0, "ordenes": []}
    real_after = real_after_stats[0] if real_after_stats else {"errores": 0, "ordenes": []}
    impacto_real = {
        "errores_before": real_before["errores"],
        "errores_after": real_after["errores"],
        "ordenes_before": len(real_before["ordenes"]),
        "ordenes_after": len(real_after["ordenes"])
    }
    return {
        "before": before,
        "after": after,
        "impacto_estimado": impacto,
        "impacto_real": impacto_real,
        "por_servicio": service_stats,
        "por_estado": status_stats,
        "periodo": {"desde": periodo_desde, "hasta": periodo_hasta},
        "filtros": {"service_type": service_type, "status": status}
    }

@api_router.post("/ai/propuestas/{propuesta_id}/accion")
async def act_on_proposal(propuesta_id: str, data: ProposalActionRequest, current_user: dict = Depends(get_current_user)):
    require_admin(current_user)
    propuesta = await db.propuestas_ia.find_one({"id": propuesta_id}, {"_id": 0})
    if not propuesta:
        raise HTTPException(status_code=404, detail="Propuesta no encontrada")
    accion = data.accion
    if accion not in ["aceptar", "rechazar", "modificar", "posponer"]:
        raise HTTPException(status_code=400, detail="Acción inválida")
    now = datetime.now(timezone.utc).isoformat()
    estado = "pendiente"
    if accion == "aceptar":
        estado = "aceptada"
    if accion == "rechazar":
        estado = "rechazada"
    if accion == "modificar":
        estado = "modificada"
    if accion == "posponer":
        estado = "pospuesta"
    await db.propuestas_ia.update_one(
        {"id": propuesta_id},
        {"$set": {"estado": estado, "feedback": {"ultima_decision": estado, "fecha_decision": now, "usuario_id": current_user["id"], "comentarios": data.comentarios}}}
    )
    feedback = {
        "id": str(uuid.uuid4()),
        "propuesta_id": propuesta_id,
        "accion_tomada": estado,
        "modificaciones": data.modificaciones,
        "motivo": data.comentarios,
        "usuario_id": current_user["id"],
        "timestamp": now
    }
    await db.feedback_ia.insert_one(feedback)
    action_payload = data.modificaciones.get("accion_sugerida") if data.modificaciones else None
    action_payload = action_payload or propuesta.get("accion_sugerida")
    if accion in ["aceptar", "modificar"] and action_payload:
        action_type = action_payload.get("type") or action_payload.get("tipo")
        payload = action_payload.get("payload", {})
        if action_type == "ajuste_regla":
            await db.reglas_negocio.update_one({"id": "order_rules_v1"}, {"$set": {"updated_at": now, **payload}}, upsert=True)
        if action_type == "nueva_validacion":
            await db.reglas_negocio.update_one({"id": "order_rules_v1"}, {"$addToSet": {"validaciones": payload}, "$set": {"updated_at": now}}, upsert=True)
        if action_type == "optimizacion_notificacion":
            await db.reglas_negocio.update_one({"id": "order_rules_v1"}, {"$set": {"updated_at": now, "auto_transitions": payload}}, upsert=True)
        if action_type == "plan_recuperacion":
            acciones = payload.get("acciones", [])
            for item in acciones:
                order_id = item.get("order_id")
                status_value = item.get("status")
                if not order_id or not status_value:
                    continue
                await db.orders.update_one(
                    {"id": order_id},
                    {
                        "$set": {
                            "status": status_value,
                            "estado_actual": status_value,
                            "updated_at": now,
                            "tiempos.ultimo_cambio_estado": now,
                            f"tiempos.fechas_estado.{status_value}": now
                        }
                    }
                )
    await create_audit_log("AI_PROPOSAL_ACTION", "propuesta_ia", propuesta_id, current_user["id"], {"accion": estado})
    return {"ok": True, "estado": estado}

@api_router.post("/admin/import")
async def create_import(origen: str = Query("csv"), file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    require_admin(current_user)
    await ensure_ai_indexes()
    filename = file.filename or ""
    content = await file.read()
    rows = []
    headers = []
    if filename.lower().endswith(".csv") or origen == "csv":
        text = content.decode("utf-8", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        headers = reader.fieldnames or []
        for idx, row in enumerate(reader):
            if idx >= 500:
                break
            rows.append(row)
    elif filename.lower().endswith(".xlsx") or origen == "excel":
        try:
            import openpyxl
        except Exception:
            raise HTTPException(status_code=400, detail="Excel no soportado sin openpyxl")
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        sheet = workbook.active
        headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        for idx, row_cells in enumerate(sheet.iter_rows(min_row=2), start=0):
            if idx >= 500:
                break
            row = {}
            for col_idx, cell in enumerate(row_cells):
                key = headers[col_idx] if col_idx < len(headers) else f"col_{col_idx}"
                row[key] = cell.value
            rows.append(row)
    else:
        raise HTTPException(status_code=400, detail="Formato no soportado")
    import_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "id": import_id,
        "origen": origen,
        "estado": "subido",
        "raw_headers": headers,
        "raw_sample": rows[:5],
        "raw_rows": rows,
        "fecha_importacion": now,
        "usuario_id": current_user["id"]
    }
    await db.importaciones_legacy.insert_one(doc)
    return {"import_id": import_id, "campos_detectados": headers, "sample": rows[:5]}

@api_router.post("/admin/import/{import_id}/mapping/suggest")
async def suggest_import_mapping(import_id: str, data: ImportMappingSuggestRequest, current_user: dict = Depends(get_current_user)):
    require_admin(current_user)
    import_doc = await db.importaciones_legacy.find_one({"id": import_id}, {"_id": 0})
    if not import_doc:
        raise HTTPException(status_code=404, detail="Importación no encontrada")
    mapping = suggest_mapping(data.campos_legacy)
    prompt = (
        "Devuelve JSON con clave sugerencias (objeto) para mapear campos legacy a ordenes. "
        f"campos={json.dumps(data.campos_legacy, ensure_ascii=False)}"
    )
    try:
        raw = call_ollama(prompt)
        payload = extract_json_payload(raw)
        ai_mapping = payload.get("sugerencias")
        if isinstance(ai_mapping, dict):
            mapping = {**mapping, **ai_mapping}
    except Exception:
        pass
    await db.importaciones_legacy.update_one({"id": import_id}, {"$set": {"mapping_sugerido": mapping}})
    return {"sugerencias": mapping}

@api_router.post("/admin/import/{import_id}/mapping/confirm")
async def confirm_import_mapping(import_id: str, data: ImportMappingConfirmRequest, current_user: dict = Depends(get_current_user)):
    require_admin(current_user)
    import_doc = await db.importaciones_legacy.find_one({"id": import_id}, {"_id": 0})
    if not import_doc:
        raise HTTPException(status_code=404, detail="Importación no encontrada")
    rows = import_doc.get("raw_rows", [])
    mapping = data.mapping_campos
    now = datetime.now(timezone.utc).isoformat()
    created_orders = []
    for row in rows:
        mapped = {}
        for legacy_key, target_path in mapping.items():
            if legacy_key in row and target_path:
                set_nested_value(mapped, target_path, row[legacy_key])
        status_value = mapped.get("estado_actual") or mapped.get("status") or "new"
        order_id = str(uuid.uuid4())
        order_number = mapped.get("order_number") or await generate_order_number()
        customer_id = await resolve_or_create_customer_from_row(mapped)
        order = {
            "id": order_id,
            "order_number": order_number,
            "customer_id": customer_id,
            "customer_name": mapped.get("customer_name"),
            "service_type": mapped.get("service_type") or "pickup_delivery",
            "pickup_date": mapped.get("formulario", {}).get("pickup_date"),
            "pickup_time_window": mapped.get("formulario", {}).get("pickup_time_window"),
            "pickup_address": mapped.get("formulario", {}).get("pickup_address"),
            "delivery_address": mapped.get("formulario", {}).get("delivery_address"),
            "estimated_lbs": mapped.get("formulario", {}).get("estimated_lbs"),
            "notes": mapped.get("notes"),
            "gate_code": mapped.get("formulario", {}).get("gate_code"),
            "status": status_value,
            "estado_actual": status_value,
            "payment_status": "unpaid",
            "total_amount": None,
            "tiempos": build_order_times(mapped.get("tiempos", {}).get("creacion") or now, status_value),
            "errores_validacion": [],
            "secciones": [],
            "importada": True,
            "origen": "import_legacy",
            "qr_token": str(uuid.uuid4()),
            "created_at": mapped.get("tiempos", {}).get("creacion") or now,
            "updated_at": now
        }
        await db.orders.insert_one(order)
        created_orders.append(order_id)
    await db.importaciones_legacy.update_one(
        {"id": import_id},
        {"$set": {"estado": "procesado", "mapping_campos": mapping, "ordenes_creadas_ids": created_orders}}
    )
    return {"ok": True, "ordenes_creadas": len(created_orders)}

@api_router.post("/admin/import/{import_id}/plan-recuperacion")
async def import_recovery_plan(import_id: str, current_user: dict = Depends(get_current_user)):
    require_admin(current_user)
    import_doc = await db.importaciones_legacy.find_one({"id": import_id}, {"_id": 0})
    if not import_doc:
        raise HTTPException(status_code=404, detail="Importación no encontrada")
    order_ids = import_doc.get("ordenes_creadas_ids", [])
    if not order_ids:
        raise HTTPException(status_code=400, detail="No hay órdenes importadas")
    now = datetime.now(timezone.utc).isoformat()
    stale_orders = await db.orders.find(
        {"id": {"$in": order_ids}, "status": {"$nin": ["completed", "cancelled"]}},
        {"_id": 0}
    ).to_list(200)
    acciones = [{"order_id": o["id"], "status": "processing"} for o in stale_orders[:50]]
    propuesta = {
        "id": str(uuid.uuid4()),
        "tipo": "plan_recuperacion",
        "descripcion": "Plan de recuperación generado desde importación legacy",
        "estado": "pendiente",
        "impacto_estimado": {"ordenes_afectadas": len(stale_orders)},
        "accion_sugerida": {"type": "plan_recuperacion", "payload": {"acciones": acciones}},
        "nivel_riesgo": "medio",
        "datos_respaldo": {"import_id": import_id},
        "fecha_generacion": now,
        "fuente": "importaciones_legacy"
    }
    await db.propuestas_ia.insert_one(propuesta)
    await db.importaciones_legacy.update_one({"id": import_id}, {"$set": {"plan_recuperacion_propuesta_id": propuesta["id"]}})
    return {"ok": True, "propuesta_id": propuesta["id"]}

@api_router.get("/public/membership-section", response_model=MembershipSectionResponse)
async def get_public_membership_section():
    section = await db.membership_section.find_one({"id": "default"}, {"_id": 0})
    if not section:
        section = {
            "id": "default",
            "heading": "Flexible Plans for Every Home",
            "subheading": None,
            "special_title": "🎉 New Member Special",
            "special_text": "$10 OFF your first month on any membership. Ask when you call or text.",
            "cta_title": "Need help choosing?",
            "cta_text": "Just call, text, or email us at (805) 836-8872 and we'll recommend the perfect plan based on your weekly laundry.",
            "cta_button_label": "👉 BECOME A MEMBER",
            "cta_button_url": "/membership",
            "contact_phone": "(805) 836-8872",
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.membership_section.insert_one(section)
    return MembershipSectionResponse(**section)

@api_router.get("/public/membership-plans", response_model=List[MembershipPlanResponse])
async def get_public_membership_plans(active_only: bool = True):
    query = {}
    if active_only:
        query["is_active"] = True
    plans = await db.membership_plans.find(query, {"_id": 0}).sort([("sort_order", 1), ("created_at", -1)]).to_list(200)
    if len(plans) == 0:
        now = datetime.now(timezone.utc).isoformat()
        seed = [
            {
                "id": str(uuid.uuid4()),
                "name": "MOST POPULAR",
                "price": "$139 / month",
                "image_url": "https://images.squarespace-cdn.com/content/v1/696c559a4b2b9b1b0febf8d7/4a2815a1-54c1-45fb-8320-244dce8b83c8/MOST+POPULAR.png",
                "features": ["Up to 60 lb/ month", "Basic Preferences saved (folding notes)", "Best value for most families"],
                "is_popular": True,
                "is_active": True,
                "sort_order": 1,
                "created_at": now,
                "updated_at": now
            },
            {
                "id": str(uuid.uuid4()),
                "name": "FAMILY PLUS",
                "price": "$199 / month",
                "image_url": "https://images.squarespace-cdn.com/content/v1/696c559a4b2b9b1b0febf8d7/f262a5b8-0043-4977-9d32-d6b343be3e70/FAMILY+PLUS.png",
                "features": ["Up to 90 lb/ month", "Priority scheduling", "Great for larger households or rentals"],
                "is_popular": False,
                "is_active": True,
                "sort_order": 2,
                "created_at": now,
                "updated_at": now
            },
            {
                "id": str(uuid.uuid4()),
                "name": "ELITE CONCIERGE",
                "price": "$299 / month",
                "image_url": "https://images.squarespace-cdn.com/content/v1/696c559a4b2b9b1b0febf8d7/13a4c501-7792-4f72-bf5c-072f95b5f995/ELITE+CONCIERGE.png",
                "features": ["Up to 120 lb/ month", "Priority turnaround (when possible)", "Premium packaging", "Saved preferences", "1 emergency pickup included"],
                "is_popular": False,
                "is_active": True,
                "sort_order": 3,
                "created_at": now,
                "updated_at": now
            }
        ]
        await db.membership_plans.insert_many(seed)
        plans = seed
    return [MembershipPlanResponse(**p) for p in plans]

# ==================== INGEST & ROUTING ====================

@api_router.post("/ingest")
async def process_ingest(data: IngestCreate, current_user: dict = Depends(get_current_user)):
    """
    Process incoming form submission and route to appropriate collection
    """
    ingest_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    # Save raw ingest
    ingest_record = {
        "id": ingest_id,
        "source_form": data.source_form,
        "data": data.data,
        "processed_flag": "locked",
        "processed_at": now,
        "route_result": None,
        "error_notes": None,
        "created_at": now
    }
    await db.ingest.insert_one(ingest_record)
    await create_audit_log("INGEST_LOCKED", "ingest", ingest_id, current_user["id"])
    
    # Route based on source_form
    route_result = "unknown"
    entity_id = None
    
    try:
        source = data.source_form.lower()
        form_data = data.data
        
        if any(kw in source for kw in ["pickup", "order", "delivery"]):
            # Route to Orders
            customer = await db.customers.find_one({"email": form_data.get("email", "").lower()}, {"_id": 0})
            if not customer:
                customer_id = str(uuid.uuid4())
                customer = {
                    "id": customer_id,
                    "name": form_data.get("name", "Unknown"),
                    "email": form_data.get("email", "").lower() or None,
                    "phone": form_data.get("phone"),
                    "address": form_data.get("address"),
                    "preferred_contact": "email",
                    "notes": None,
                    "status": "active",
                    "total_orders": 0,
                    "created_at": now,
                    "updated_at": now
                }
                await db.customers.insert_one(customer)
            
            order_id = str(uuid.uuid4())
            order_number = await generate_order_number()
            order = {
                "id": order_id,
                "order_number": order_number,
                "customer_id": customer["id"],
                "customer_name": customer["name"],
                "service_type": form_data.get("service_type", "pickup_delivery"),
                "pickup_date": form_data.get("pickup_date"),
                "pickup_time_window": form_data.get("pickup_time"),
                "pickup_address": form_data.get("address"),
                "delivery_address": form_data.get("delivery_address"),
                "estimated_lbs": form_data.get("estimated_lbs"),
                "actual_lbs": None,
                "notes": form_data.get("notes"),
                "gate_code": form_data.get("gate_code"),
                "status": "new",
                "payment_status": "unpaid",
                "total_amount": None,
                "created_at": now,
                "updated_at": now
            }
            await db.orders.insert_one(order)
            await db.customers.update_one({"id": customer["id"]}, {"$inc": {"total_orders": 1}})
            route_result = "orders"
            entity_id = order_id
            
        elif any(kw in source for kw in ["quote", "commercial", "b2b", "business"]):
            # Route to Quotes
            quote_id = str(uuid.uuid4())
            quote_number = await generate_quote_number()
            quote = {
                "id": quote_id,
                "quote_number": quote_number,
                "company_name": form_data.get("company_name", form_data.get("name", "Unknown")),
                "contact_name": form_data.get("contact_name", form_data.get("name", "")),
                "email": form_data.get("email", "").lower() or None,
                "phone": form_data.get("phone"),
                "industry": form_data.get("industry"),
                "estimated_lbs_per_week": form_data.get("estimated_lbs"),
                "service_needs": form_data.get("service_needs"),
                "notes": form_data.get("notes"),
                "status": "new",
                "assigned_to": None,
                "follow_up_date": None,
                "created_at": now,
                "updated_at": now
            }
            await db.quotes.insert_one(quote)
            route_result = "quotes"
            entity_id = quote_id
            
        elif any(kw in source for kw in ["support", "ticket", "issue", "feedback", "complaint"]):
            # Route to Support Tickets
            ticket_id = str(uuid.uuid4())
            ticket_number = await generate_ticket_number()
            subject = form_data.get("subject", form_data.get("regarding", "Support Request"))
            description = form_data.get("description", form_data.get("message", ""))
            priority = determine_priority(subject, description)
            
            ticket = {
                "id": ticket_id,
                "ticket_number": ticket_number,
                "customer_id": None,
                "customer_name": form_data.get("name"),
                "subject": subject,
                "description": description,
                "category": form_data.get("category", "general"),
                "priority": priority,
                "status": "open",
                "assigned_to": None,
                "resolution": None,
                "created_at": now,
                "updated_at": now
            }
            await db.tickets.insert_one(ticket)
            route_result = "tickets"
            entity_id = ticket_id
            
        else:
            # Default to Leads
            lead_id = str(uuid.uuid4())
            lead = {
                "id": lead_id,
                "name": form_data.get("name", "Unknown"),
                "email": form_data.get("email", "").lower() or None,
                "phone": form_data.get("phone"),
                "source": data.source_form,
                "interest_type": form_data.get("interest_type", form_data.get("service_type")),
                "notes": form_data.get("notes", form_data.get("message")),
                "status": "new",
                "converted_to_customer_id": None,
                "created_at": now,
                "updated_at": now
            }
            await db.leads.insert_one(lead)
            route_result = "leads"
            entity_id = lead_id
        
        # Update ingest record
        await db.ingest.update_one(
            {"id": ingest_id},
            {"$set": {"processed_flag": "processed", "route_result": route_result, "route_id": entity_id}}
        )
        await create_audit_log("INGEST_ROUTED", "ingest", ingest_id, current_user["id"], {"route": route_result, "entity_id": entity_id})
        
    except Exception as e:
        await db.ingest.update_one(
            {"id": ingest_id},
            {"$set": {"processed_flag": "error", "error_notes": str(e)}}
        )
        await create_audit_log("INGEST_ERROR", "ingest", ingest_id, current_user["id"], {"error": str(e)})
        raise HTTPException(status_code=500, detail=f"Error processing ingest: {str(e)}")
    
    return {
        "ingest_id": ingest_id,
        "route_result": route_result,
        "entity_id": entity_id,
        "message": f"Form submission routed to {route_result}"
    }

@api_router.get("/ingest", response_model=List[dict])
async def get_ingest_records(
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query["processed_flag"] = status
    
    records = await db.ingest.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return records

# ==================== AUDIT LOG ====================

@api_router.get("/audit-logs", response_model=List[AuditLogResponse])
async def get_audit_logs(
    entity_type: Optional[str] = None,
    entity_id: Optional[str] = None,
    event_type: Optional[str] = None,
    limit: int = 100,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if entity_type:
        query["entity_type"] = entity_type
    if entity_id:
        query["entity_id"] = entity_id
    if event_type:
        query["event_type"] = event_type
    
    logs = await db.audit_logs.find(query, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
    return [AuditLogResponse(**log) for log in logs]

# ==================== HEALTH CHECK ====================

@api_router.get("/")
async def root():
    return {"message": "Ventura Fresh Laundry CRM API", "status": "healthy"}

@api_router.get("/health")
async def health_check():
    return {"status": "healthy", "timestamp": datetime.now(timezone.utc).isoformat()}

# ==================== EXPORT ENDPOINTS ====================

@api_router.get("/export/customers")
async def export_customers_csv(current_user: dict = Depends(get_current_user)):
    """Export customers to CSV"""
    customers = await db.customers.find({}, {"_id": 0}).to_list(10000)

    output = io.StringIO()
    if customers:
        all_keys = set()
        for customer in customers:
            if isinstance(customer, dict):
                all_keys.update([str(key) for key in customer.keys()])
        fieldnames = sorted(list(all_keys))

        writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for customer in customers:
            normalized = {}
            for key in fieldnames:
                value = customer.get(key, "")
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, default=str)
                elif value is None:
                    value = ""
                elif not isinstance(value, (str, int, float, bool)):
                    value = str(value)
                normalized[key] = value
            writer.writerow(normalized)

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=customers.csv"}
    )

@api_router.get("/export/orders")
async def export_orders_csv(current_user: dict = Depends(get_current_user)):
    """Export orders to CSV"""
    orders = await db.orders.find({}, {"_id": 0}).to_list(10000)
    
    output = io.StringIO()
    if orders:
        base_fields = [
            "id",
            "order_number",
            "customer_id",
            "customer_name",
            "service_type",
            "pickup_date",
            "pickup_time_window",
            "pickup_address",
            "delivery_address",
            "estimated_lbs",
            "notes",
            "gate_code",
            "status",
            "payment_status",
            "total_amount",
            "created_at",
            "updated_at"
        ]
        fieldnames = [field for field in base_fields if any(field in o for o in orders)]
        extra_fields = sorted({key for o in orders for key in o.keys()} - set(fieldnames))
        fieldnames.extend(extra_fields)
        writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(orders)
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=orders.csv"}
    )

@api_router.get("/export/leads")
async def export_leads_csv(current_user: dict = Depends(get_current_user)):
    """Export leads to CSV"""
    leads = await db.leads.find({}, {"_id": 0}).to_list(10000)
    
    output = io.StringIO()
    if leads:
        writer = csv.DictWriter(output, fieldnames=leads[0].keys())
        writer.writeheader()
        writer.writerows(leads)
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=leads.csv"}
    )

@api_router.get("/export/quotes")
async def export_quotes_csv(current_user: dict = Depends(get_current_user)):
    """Export quotes to CSV"""
    quotes = await db.quotes.find({}, {"_id": 0}).to_list(10000)
    
    output = io.StringIO()
    if quotes:
        writer = csv.DictWriter(output, fieldnames=quotes[0].keys())
        writer.writeheader()
        writer.writerows(quotes)
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=quotes.csv"}
    )

@api_router.get("/export/tickets")
async def export_tickets_csv(current_user: dict = Depends(get_current_user)):
    """Export tickets to CSV"""
    tickets = await db.tickets.find({}, {"_id": 0}).to_list(10000)
    
    output = io.StringIO()
    if tickets:
        writer = csv.DictWriter(output, fieldnames=tickets[0].keys())
        writer.writeheader()
        writer.writerows(tickets)
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=tickets.csv"}
    )

# ==================== CALENDAR ENDPOINTS ====================

@api_router.get("/calendar/orders")
async def get_calendar_orders(
    start_date: str = Query(..., description="Start date YYYY-MM-DD"),
    end_date: str = Query(..., description="End date YYYY-MM-DD"),
    current_user: dict = Depends(get_current_user)
):
    """Get orders for calendar view within date range"""
    orders = await db.orders.find({
        "pickup_date": {"$gte": start_date, "$lte": end_date}
    }, {"_id": 0}).to_list(1000)
    
    # Format for calendar
    events = []
    for order in orders:
        events.append({
            "id": order["id"],
            "title": f"{order['order_number']} - {order['customer_name']}",
            "date": order["pickup_date"],
            "time": order.get("pickup_time_window"),
            "status": order["status"],
            "service_type": order["service_type"],
            "address": order.get("pickup_address"),
            "customer_name": order["customer_name"],
            "order_number": order["order_number"]
        })
    
    return events

# ==================== NOTIFICATION SETTINGS ====================

@api_router.get("/settings/notifications")
async def get_notification_settings(current_user: dict = Depends(get_current_user)):
    """Get notification service status"""
    return {
        "email_enabled": bool(os.environ.get('SENDGRID_API_KEY') and os.environ.get('SENDGRID_FROM_EMAIL')),
        "sms_enabled": bool(os.environ.get('TWILIO_ACCOUNT_SID') and os.environ.get('TWILIO_AUTH_TOKEN')),
        "voice_enabled": bool(os.environ.get('TWILIO_ACCOUNT_SID') and os.environ.get('TWILIO_AUTH_TOKEN') and os.environ.get('TWILIO_PHONE_NUMBER')),
        "notifications_available": NOTIFICATIONS_ENABLED
    }

@api_router.get("/settings/rules")
async def get_business_rules(current_user: dict = Depends(get_current_user)):
    require_admin(current_user)
    rules = await get_or_seed_business_rules()
    return rules

@api_router.put("/settings/rules")
async def update_business_rules(data: RulesUpdateRequest, current_user: dict = Depends(get_current_user)):
    require_admin(current_user)
    rules = data.rules or {}
    now = datetime.now(timezone.utc).isoformat()
    rules["updated_at"] = now
    rules.setdefault("id", "order_rules_v1")
    await db.reglas_negocio.update_one({"id": rules["id"]}, {"$set": rules}, upsert=True)
    await create_audit_log("RULES_UPDATED", "reglas_negocio", rules["id"], current_user["id"])
    return rules

@api_router.post("/test/email")
async def test_email_notification(
    to_email: EmailStr,
    current_user: dict = Depends(get_current_user)
):
    """Test email notification"""
    if not NOTIFICATIONS_ENABLED:
        raise HTTPException(status_code=400, detail="Notifications not configured")
    
    result = await send_email(
        to_email,
        "Test - Ventura Fresh Laundry CRM",
        "<h1>Test Email</h1><p>Este es un correo de prueba del CRM.</p>"
    )
    return result

@api_router.post("/test/sms")
async def test_sms_notification(
    to_phone: str,
    current_user: dict = Depends(get_current_user)
):
    """Test SMS notification"""
    if not NOTIFICATIONS_ENABLED:
        raise HTTPException(status_code=400, detail="Notifications not configured")
    
    result = await send_sms(to_phone, "Test: Este es un mensaje de prueba del CRM de Ventura Fresh Laundry.")
    return result

# ==================== CUSTOMER AUTHENTICATION ====================

class CustomerRegister(BaseModel):
    name: str
    email: EmailStr
    password: str

class CustomerLogin(BaseModel):
    email: EmailStr
    password: str

class CustomerAuthResponse(BaseModel):
    access_token: str
    token_type: str
    customer: dict

@api_router.post("/customer/auth/register", response_model=CustomerAuthResponse)
async def customer_register(data: CustomerRegister):
    """Register a new customer account"""
    existing = await db.customers.find_one({"email": data.email.lower()})
    
    if existing:
        # Check if customer already has a password (already registered)
        if existing.get("password_hash"):
            raise HTTPException(status_code=400, detail="Email already registered. Please login.")
        
        # Customer exists from a pickup request but hasn't registered - update with password
        await db.customers.update_one(
            {"email": data.email.lower()},
            {"$set": {
                "password_hash": hash_password(data.password),
                "name": data.name,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        customer = await db.customers.find_one({"email": data.email.lower()}, {"_id": 0, "password_hash": 0})
    else:
        # Create new customer
        customer_id = str(uuid.uuid4())
        now = datetime.now(timezone.utc).isoformat()
        customer = {
            "id": customer_id,
            "name": data.name,
            "email": data.email.lower(),
            "phone": None,
            "address": None,
            "preferred_contact": "email",
            "notes": None,
            "status": "active",
            "total_orders": 0,
            "password_hash": hash_password(data.password),
            "created_at": now,
            "updated_at": now
        }
        await db.customers.insert_one(customer)
        await create_audit_log("CUSTOMER_REGISTERED", "customer", customer_id, None, {"source": "portal"})
        customer = {k: v for k, v in customer.items() if k not in ["password_hash", "_id"]}
    
    token = create_customer_token(customer["id"], customer["email"])
    return CustomerAuthResponse(
        access_token=token,
        token_type="bearer",
        customer=customer
    )

@api_router.post("/customer/auth/login", response_model=CustomerAuthResponse)
async def customer_login(data: CustomerLogin):
    """Customer login"""
    customer = await db.customers.find_one({"email": data.email.lower()})
    if not customer:
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    if not customer.get("password_hash"):
        raise HTTPException(status_code=401, detail="Please register an account first")
    
    if not verify_password(data.password, customer["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    customer_data = {k: v for k, v in customer.items() if k not in ["_id", "password_hash"]}
    token = create_customer_token(customer["id"], customer["email"])
    
    return CustomerAuthResponse(
        access_token=token,
        token_type="bearer",
        customer=customer_data
    )

@api_router.get("/customer/me")
async def get_customer_profile(current_customer: dict = Depends(get_current_customer)):
    """Get current customer profile"""
    return current_customer

@api_router.get("/customer/orders")
async def get_customer_orders(current_customer: dict = Depends(get_current_customer)):
    """Get orders for the logged-in customer"""
    orders = await db.orders.find(
        {"customer_id": current_customer["id"]}, 
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return orders

# ==================== USER MANAGEMENT (ADMIN ONLY) ====================

class UserUpdateRole(BaseModel):
    role: str

class UserCreateAdmin(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: str = ROLE_OPERATOR

@api_router.get("/admin/users")
async def list_users(current_user: dict = Depends(get_current_user)):
    """List all users (admin only)"""
    require_admin(current_user)
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).sort("created_at", -1).to_list(100)
    return users

@api_router.post("/admin/users", response_model=UserResponse)
async def create_user_admin(user_data: UserCreateAdmin, current_user: dict = Depends(get_current_user)):
    """Create a new user with specified role (admin only)"""
    require_admin(current_user)
    
    if user_data.role not in VALID_ROLES:
        raise HTTPException(status_code=400, detail=f"Invalid role. Must be one of: {VALID_ROLES}")
    
    existing = await db.users.find_one({"email": user_data.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    user = {
        "id": user_id,
        "email": user_data.email.lower(),
        "password_hash": hash_password(user_data.password),
        "name": user_data.name,
        "role": user_data.role,
        "created_at": now
    }
    await db.users.insert_one(user)
    await create_audit_log("USER_CREATED_BY_ADMIN", "user", user_id, current_user["id"], {"role": user_data.role})
    
    return UserResponse(id=user_id, email=user["email"], name=user["name"], role=user["role"], created_at=now)

@api_router.put("/admin/users/{user_id}/role")
async def update_user_role(user_id: str, data: UserUpdateRole, current_user: dict = Depends(get_current_user)):
    """Update user role (admin only)"""
    require_admin(current_user)
    
    if data.role not in VALID_ROLES:
        raise HTTPException(status_code=400, detail=f"Invalid role. Must be one of: {VALID_ROLES}")
    
    # Prevent admin from demoting themselves
    if user_id == current_user["id"] and data.role != ROLE_ADMIN:
        raise HTTPException(status_code=400, detail="Cannot change your own role")
    
    result = await db.users.update_one(
        {"id": user_id},
        {"$set": {"role": data.role}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    await create_audit_log("USER_ROLE_UPDATED", "user", user_id, current_user["id"], {"new_role": data.role})
    return {"message": f"User role updated to {data.role}"}

@api_router.delete("/admin/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a user (admin only)"""
    require_admin(current_user)
    
    # Prevent admin from deleting themselves
    if user_id == current_user["id"]:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    await create_audit_log("USER_DELETED", "user", user_id, current_user["id"])
    return {"message": "User deleted"}

@api_router.get("/admin/roles")
async def get_roles(current_user: dict = Depends(get_current_user)):
    """Get available roles and their permissions"""
    require_admin(current_user)
    return {
        "roles": VALID_ROLES,
        "permissions": ROLE_PERMISSIONS
    }

# ==================== OPERATOR-ONLY ENDPOINTS ====================

@api_router.get("/operator/orders")
async def operator_get_orders(
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get orders for operator (limited view - no financial data)"""
    if not has_permission(current_user, "orders:read"):
        raise HTTPException(status_code=403, detail="Permission denied")
    
    query = {}
    if status:
        query["status"] = status
    
    orders = await db.orders.find(query, {
        "_id": 0,
        "total_amount": 0,  # Hide financial data from operators
        "payment_status": 0
    }).sort("created_at", -1).to_list(500)
    return orders

@api_router.patch("/operator/orders/{order_id}/status")
async def operator_update_order_status(
    order_id: str, 
    status: str, 
    current_user: dict = Depends(get_current_user)
):
    """Update order status (operator allowed)"""
    if not has_permission(current_user, "orders:update_status"):
        raise HTTPException(status_code=403, detail="Permission denied")
    
    valid_statuses = ["new", "confirmed", "pickup_scheduled", "picked_up", "processing", "ready", "out_for_delivery", "delivered", "completed", "cancelled"]
    status_normalized = normalize_status(status)
    if status_normalized not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {valid_statuses}")
    
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    current_status = normalize_status(order.get("status"))
    service_type = normalize_spaces(order.get("service_type") or "pickup_delivery").lower().replace(" ", "_")
    if status_normalized == "completed":
        if service_type in ["wash_fold", "wash_fold_dropoff", "self_service"]:
            if current_status not in ["ready", "completed"]:
                raise HTTPException(status_code=400, detail="Wash & Fold must be ready before it can be completed")
        elif current_status not in ["delivered", "completed", "out_for_delivery"]:
            raise HTTPException(status_code=400, detail="Order must be delivered before it can be completed")

    now = datetime.now(timezone.utc).isoformat()
    result = await db.orders.update_one(
        {"id": order_id},
        {
            "$set": {
                "status": status_normalized,
                "estado_actual": status_normalized,
                "updated_at": now,
                "tiempos.ultimo_cambio_estado": now,
                f"tiempos.fechas_estado.{status_normalized}": now
            }
        }
    )
    
    await create_audit_log("ORDER_STATUS_CHANGED_BY_OPERATOR", "order", order_id, current_user["id"], {"new_status": status_normalized})
    
    # Send notifications if enabled
    if NOTIFICATIONS_ENABLED and not SKIP_SERVER_NOTIFICATIONS and order.get("customer_id"):
        customer = await db.customers.find_one({"id": order["customer_id"]}, {"_id": 0})
        if customer and should_notify_order_status(order, status_normalized):
            order["status"] = status_normalized
            try:
                await notify_order_status_changed(customer, order, status_normalized)
            except Exception as e:
                logger.error(f"Notification failed: {e}")
    
    return {"message": f"Order status updated to {status_normalized}"}

# === External routers (refactored) ===
if get_public_forms_router:
    public_forms_router = get_public_forms_router(
        db=db,
        generate_order_number=generate_order_number,
        create_audit_log=create_audit_log,
        emit_realtime=emit_realtime,
        notifications_enabled=NOTIFICATIONS_ENABLED,
        skip_server_notifications=SKIP_SERVER_NOTIFICATIONS,
        logger=logger
    )
    api_router.include_router(public_forms_router)

if NOTIFICATIONS_ENABLED and get_voice_router:
    voice_router = get_voice_router(
        db=db,
        require_admin=require_admin,
        get_current_user=get_current_user,
        build_notification_content=build_notification_content,
        send_voice_call=send_voice_call,
        detect_language=detect_language,
        generate_ai_message=generate_ai_message,
        normalize_phone=normalize_phone,
        create_audit_log=create_audit_log
    )
    api_router.include_router(voice_router)

# Include router and middleware
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Include n8n router
if N8N_ENABLED and n8n_router:
    app.include_router(n8n_router, prefix="/api")
    logger.info("n8n integration endpoints enabled at /api/n8n/*")

# Include store router
if STORE_ENABLED and store_router:
    app.include_router(store_router, prefix="/api")
    logger.info("Store endpoints enabled at /api/store/*")

# Include blog router
if BLOG_ENABLED and blog_router:
    app.include_router(blog_router, prefix="/api")
    logger.info("Blog endpoints enabled at /api/blog/*")

# Include automation engine router
if AUTOMATION_ENABLED and automation_router:
    app.include_router(automation_router, prefix="/api")
    logger.info("Automation engine enabled at /api/automation/*")

# Include Stripe advanced sync scaffold (feature-flag controlled inside module)
if STRIPE_SYNC_SCAFFOLD_ENABLED and stripe_sync_router:
    app.include_router(stripe_sync_router, prefix="/api")
    logger.info("Stripe sync scaffold enabled at /api/stripe-sync/*")

# Stripe webhook endpoint
@app.post("/api/webhook/stripe")
async def stripe_webhook(request: Request):
    """Handle Stripe webhook events"""
    if STORE_ENABLED:
        return await handle_stripe_webhook(request)
    raise HTTPException(status_code=503, detail="Store module not available")

# ==================== STATIC WEBSITE ROUTES ====================
# Serve the HTML website files

WEB_DIR = ROOT_DIR / "paginaweb"
UPLOADS_DIR = ROOT_DIR / "uploads"
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)

app.mount("/uploads", StaticFiles(directory=str(UPLOADS_DIR)), name="uploads")
# Mount static files directories for each page's assets
if WEB_DIR.exists():
    for folder in WEB_DIR.iterdir():
        if folder.is_dir() and (folder.name.endswith('_files') or folder.name.endswith('_resources')):
            app.mount(f"/web/{folder.name}", StaticFiles(directory=folder), name=folder.name)


@app.get("/web/crm-integration.js")
async def serve_crm_js():
    """Serve the CRM integration JavaScript"""
    js_file = WEB_DIR / "crm-integration.js"
    if js_file.exists():
        return FileResponse(js_file, media_type="application/javascript")
    raise HTTPException(status_code=404, detail="File not found")

@app.get("/web/", response_class=HTMLResponse)
@app.get("/web", response_class=HTMLResponse)
async def serve_home():
    """Serve the main landing page"""
    html_file = WEB_DIR / "index.html"
    if html_file.exists():
        return HTMLResponse(content=html_file.read_text(encoding='utf-8'))
    raise HTTPException(status_code=404, detail="Page not found")

@app.get("/web/about", response_class=HTMLResponse)
async def serve_about():
    """Serve the about page"""
    html_file = WEB_DIR / "about.html"
    if html_file.exists():
        return HTMLResponse(content=html_file.read_text(encoding='utf-8'))
    raise HTTPException(status_code=404, detail="Page not found")

@app.get("/web/contact", response_class=HTMLResponse)
async def serve_contact():
    """Serve the contact page"""
    html_file = WEB_DIR / "contact.html"
    if html_file.exists():
        return HTMLResponse(content=html_file.read_text(encoding='utf-8'))
    raise HTTPException(status_code=404, detail="Page not found")

@app.get("/web/services", response_class=HTMLResponse)
async def serve_services():
    """Serve the services page"""
    html_file = WEB_DIR / "services.html"
    if html_file.exists():
        return HTMLResponse(content=html_file.read_text(encoding='utf-8'))
    raise HTTPException(status_code=404, detail="Page not found")

@app.get("/web/blog", response_class=HTMLResponse)
async def serve_blog():
    """Serve the blog page"""
    html_file = WEB_DIR / "blog.html"
    if html_file.exists():
        return HTMLResponse(content=html_file.read_text(encoding='utf-8'))
    raise HTTPException(status_code=404, detail="Page not found")

@app.get("/web/store", response_class=HTMLResponse)
async def serve_store():
    """Serve the store page"""
    html_file = WEB_DIR / "store.html"
    if html_file.exists():
        return HTMLResponse(content=html_file.read_text(encoding='utf-8'))
    raise HTTPException(status_code=404, detail="Page not found")

@app.get("/web/schedule", response_class=HTMLResponse)
async def serve_schedule():
    """Serve the schedule pickup page"""
    html_file = WEB_DIR / "schedule.html"
    if html_file.exists():
        return HTMLResponse(content=html_file.read_text(encoding='utf-8'))
    raise HTTPException(status_code=404, detail="Page not found")

@app.get("/web/account", response_class=HTMLResponse)
async def serve_account():
    """Serve the account page"""
    html_file = WEB_DIR / "account.html"
    if html_file.exists():
        return HTMLResponse(content=html_file.read_text(encoding='utf-8'))
    raise HTTPException(status_code=404, detail="Page not found")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

# Socket.IO wrapper (served under /api/socket.io)
app = socketio.ASGIApp(sio, other_asgi_app=fastapi_app, socketio_path="api/socket.io")
